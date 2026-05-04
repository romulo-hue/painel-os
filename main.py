import os
import json
import shutil
import base64
from io import BytesIO
from typing import Any, Dict, Generator, List, Optional
from collections import defaultdict
from datetime import datetime, timedelta

import requests

from fastapi import Depends, FastAPI, HTTPException, File, UploadFile, Form, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import Column, ForeignKey, Integer, String, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, relationship
from starlette.middleware.sessions import SessionMiddleware

from database import Base, SessionLocal, engine

app = FastAPI(title="API Frota", version="5.6.0")

app.add_middleware(
    SessionMiddleware,
    secret_key="frota-secret-key-123456",
    session_cookie="frota_session",
    max_age=60 * 60 * 8,
)

UPLOAD_DIR = "uploads/fotos"
os.makedirs(UPLOAD_DIR, exist_ok=True)

app.mount("/fotos", StaticFiles(directory=UPLOAD_DIR), name="fotos")


# Endpoints da API que realmente envia ao FrotaWeb.
# Configure no Render se precisar trocar.
FROTAWEB_OS_URL = os.getenv(
    "FROTAWEB_OS_URL",
    "https://frotaweb-os-corretiva-api.onrender.com/os-corretiva?simulacao=false",
)
FROTAWEB_SERVICO_URL = os.getenv(
    "FROTAWEB_SERVICO_URL",
    "https://frotaweb-os-corretiva-api.onrender.com/os-corretiva/servicos?simulacao=false",
)



# =========================
# DATABASE
# =========================

def get_db() -> Generator[Session, None, None]:
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def usuario_logado(request: Request) -> bool:
    return request.session.get("usuario_logado", False) is True


# =========================
# MODELS
# =========================

class ServicoModel(Base):
    __tablename__ = "servicos"

    id = Column(Integer, primary_key=True, index=True)
    codigo = Column(String, nullable=True, index=True)
    nome = Column(String, unique=True, index=True, nullable=False)
    sistema = Column(String, nullable=True, index=True)


class ProdutoModel(Base):
    __tablename__ = "produtos"

    id = Column(Integer, primary_key=True, index=True)
    codigo = Column(String, nullable=True, index=True)
    nome = Column(String, unique=True, index=True, nullable=False)
    referencia = Column(String, nullable=True, index=True)
    unidade = Column(String, nullable=True)
    ativo = Column(String, nullable=False, default="S")


class MecanicoModel(Base):
    __tablename__ = "mecanicos"

    id = Column(Integer, primary_key=True, index=True)
    codigo = Column(String, unique=True, index=True, nullable=False)
    nome = Column(String, nullable=False, index=True)
    senha = Column(String, nullable=False)


class OrdemServicoModel(Base):
    __tablename__ = "ordens_servico"

    id = Column(Integer, primary_key=True, index=True)
    placa = Column(String, nullable=False, index=True)
    matricula = Column(String, nullable=True, index=True)
    mecanico = Column(String, nullable=True, index=True)
    km = Column(String, nullable=True, index=True)
    hora_inicio = Column(String, nullable=True)
    hora_fim = Column(String, nullable=True)
    observacoes = Column(String, nullable=True)
    data_cadastro = Column(String, nullable=True, index=True)
    foto = Column(String, nullable=True)

    # Controle do envio manual ao FrotaWeb
    frotaweb_status = Column(String, nullable=False, default="PENDENTE")
    frotaweb_numero_os = Column(String, nullable=True)
    frotaweb_retorno = Column(String, nullable=True)
    frotaweb_payload = Column(String, nullable=True)

    itens = relationship(
        "OrdemServicoItemModel",
        back_populates="ordem",
        cascade="all, delete-orphan",
    )

    produtos_consumidos = relationship(
        "OrdemServicoProdutoModel",
        back_populates="ordem",
        cascade="all, delete-orphan",
    )


class OrdemServicoItemModel(Base):
    __tablename__ = "ordens_servico_itens"

    id = Column(Integer, primary_key=True, index=True)
    ordem_id = Column(Integer, ForeignKey("ordens_servico.id"), nullable=False, index=True)
    servico = Column(String, nullable=False, index=True)

    # Controle do envio manual do serviço ao FrotaWeb
    frotaweb_status = Column(String, nullable=False, default="PENDENTE")
    frotaweb_retorno = Column(String, nullable=True)
    frotaweb_payload = Column(String, nullable=True)

    ordem = relationship("OrdemServicoModel", back_populates="itens")


class OrdemServicoProdutoModel(Base):
    __tablename__ = "ordens_servico_produtos"

    id = Column(Integer, primary_key=True, index=True)
    ordem_id = Column(Integer, ForeignKey("ordens_servico.id"), nullable=False, index=True)
    produto_id = Column(Integer, ForeignKey("produtos.id"), nullable=True, index=True)
    produto_nome = Column(String, nullable=False)
    quantidade = Column(String, nullable=False)

    ordem = relationship("OrdemServicoModel", back_populates="produtos_consumidos")


class VeiculoModel(Base):
    __tablename__ = "veiculos"

    id = Column(Integer, primary_key=True, index=True)
    codigo_frota = Column(String, nullable=False, index=True)
    placa = Column(String, nullable=False, unique=True, index=True)
    cd_ccusto = Column(String, nullable=False, index=True)
    cd_filial = Column(String, nullable=False, index=True)


class FrotaWebConfigModel(Base):
    __tablename__ = "frotaweb_config"

    id = Column(Integer, primary_key=True, index=True)
    matricula = Column(String, nullable=False, unique=True, index=True)
    usuario = Column(String, nullable=False)
    senha = Column(String, nullable=False)
    filial_login = Column(String, nullable=False)
    cd_empresa = Column(String, nullable=False, default="1")


# =========================
# SCHEMAS
# =========================

class ServicoCreate(BaseModel):
    codigo: str = ""
    nome: str
    sistema: str = ""


class ServicoUpdate(BaseModel):
    codigo: str = ""
    nome: str
    sistema: str = ""


class ProdutoCreate(BaseModel):
    codigo: str = ""
    nome: str
    referencia: str = ""
    unidade: str = ""
    ativo: str = "S"


class ProdutoUpdate(BaseModel):
    codigo: str = ""
    nome: str
    referencia: str = ""
    unidade: str = ""
    ativo: str = "S"


class MecanicoCreate(BaseModel):
    codigo: str
    nome: str
    senha: str


class MecanicoUpdate(BaseModel):
    codigo: str
    nome: str
    senha: str


class ProdutoConsumoCreate(BaseModel):
    produto_id: Optional[int] = None
    produto_nome: str
    quantidade: str


class OrdemServicoCreate(BaseModel):
    placa: str
    matricula: str
    mecanico: str
    km: str
    hora_inicio: str
    hora_fim: str
    servicos: List[str]
    produtos: List[ProdutoConsumoCreate] = []
    observacoes: str = ""
    data_cadastro: str


class VeiculoCreate(BaseModel):
    codigo_frota: str
    placa: str
    cd_ccusto: str
    cd_filial: str


class VeiculoUpdate(BaseModel):
    codigo_frota: str
    placa: str
    cd_ccusto: str
    cd_filial: str


class FrotaWebConfigCreate(BaseModel):
    matricula: str
    usuario: str
    senha: str
    filial_login: str
    cd_empresa: str = "1"


class FrotaWebConfigUpdate(BaseModel):
    usuario: str
    senha: str
    filial_login: str
    cd_empresa: str = "1"


class OrdemServicoFrotaWebRequest(BaseModel):
    veiculo_id: int
    matricula: str
    km: str
    hora_inicio: str
    hora_fim: str
    observacoes: str = ""
    data_cadastro: str
    cd_servicos: List[str]


# =========================
# HELPERS
# =========================

def parse_data_brasileira(data_str: str):
    try:
        return datetime.strptime(data_str, "%d/%m/%Y %H:%M")
    except Exception:
        return None


def hora_para_minutos(hora: str):
    try:
        h, m = hora.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return None


def calcular_duracao_minutos(hora_inicio: str, hora_fim: str) -> int:
    inicio = hora_para_minutos(hora_inicio or "")
    fim = hora_para_minutos(hora_fim or "")
    if inicio is None or fim is None:
        return 0
    if fim < inicio:
        return 0
    return fim - inicio


def minutos_para_texto(total_minutos: int) -> str:
    horas = total_minutos // 60
    minutos = total_minutos % 60
    return f"{horas}h {minutos:02d}min"


def data_no_intervalo(data_cadastro: str, data_inicial: str = "", data_final: str = "") -> bool:
    if not data_inicial.strip() and not data_final.strip():
        return True

    dt = parse_data_brasileira(data_cadastro)
    if dt is None:
        return False

    if data_inicial.strip():
        try:
            dt_ini = datetime.strptime(data_inicial.strip(), "%Y-%m-%d")
            if dt.date() < dt_ini.date():
                return False
        except Exception:
            pass

    if data_final.strip():
        try:
            dt_fim = datetime.strptime(data_final.strip(), "%Y-%m-%d")
            if dt.date() > dt_fim.date():
                return False
        except Exception:
            pass

    return True


def aplicar_filtros_ordens(
    query,
    placa: str = "",
    matricula: str = "",
    mecanico: str = "",
    km: str = "",
    servico: str = "",
    data_cadastro: str = "",
):
    if placa.strip():
        query = query.filter(OrdemServicoModel.placa.ilike(f"%{placa.strip()}%"))

    if matricula.strip():
        query = query.filter(OrdemServicoModel.matricula.ilike(f"%{matricula.strip()}%"))

    if mecanico.strip():
        query = query.filter(OrdemServicoModel.mecanico.ilike(f"%{mecanico.strip()}%"))

    if km.strip():
        query = query.filter(OrdemServicoModel.km.ilike(f"%{km.strip()}%"))

    if servico.strip():
        query = query.filter(OrdemServicoItemModel.servico.ilike(f"%{servico.strip()}%"))

    if data_cadastro.strip():
        query = query.filter(OrdemServicoModel.data_cadastro.ilike(f"%{data_cadastro.strip()}%"))

    return query


def buscar_linhas_ordens(
    db: Session,
    placa: str = "",
    matricula: str = "",
    mecanico: str = "",
    km: str = "",
    servico: str = "",
    data_cadastro: str = "",
):
    query = (
        db.query(
            OrdemServicoModel.id.label("id"),
            OrdemServicoModel.placa.label("placa"),
            OrdemServicoModel.matricula.label("matricula"),
            OrdemServicoModel.mecanico.label("mecanico"),
            OrdemServicoModel.km.label("km"),
            OrdemServicoModel.hora_inicio.label("hora_inicio"),
            OrdemServicoModel.hora_fim.label("hora_fim"),
            OrdemServicoItemModel.servico.label("servico"),
            OrdemServicoModel.observacoes.label("observacoes"),
            OrdemServicoModel.data_cadastro.label("data_cadastro"),
            OrdemServicoModel.foto.label("foto"),
            OrdemServicoModel.frotaweb_status.label("frotaweb_status"),
            OrdemServicoModel.frotaweb_numero_os.label("frotaweb_numero_os"),
            OrdemServicoModel.frotaweb_retorno.label("frotaweb_retorno"),
        )
        .join(OrdemServicoItemModel, OrdemServicoItemModel.ordem_id == OrdemServicoModel.id)
    )

    query = aplicar_filtros_ordens(
        query,
        placa=placa,
        matricula=matricula,
        mecanico=mecanico,
        km=km,
        servico=servico,
        data_cadastro=data_cadastro,
    )

    return query.order_by(
        OrdemServicoModel.id.desc(),
        OrdemServicoItemModel.id.asc(),
    ).all()


def buscar_ordens_dashboard(
    db: Session,
    data_inicial: str = "",
    data_final: str = "",
):
    ordens = (
        db.query(OrdemServicoModel)
        .order_by(OrdemServicoModel.id.desc())
        .all()
    )

    filtradas = []
    for ordem in ordens:
        if data_no_intervalo(ordem.data_cadastro or "", data_inicial, data_final):
            filtradas.append(ordem)

    return filtradas


def obter_intervalo_periodo_anterior(data_inicial: str = "", data_final: str = ""):
    if not data_inicial.strip() or not data_final.strip():
        return "", ""

    try:
        dt_ini = datetime.strptime(data_inicial.strip(), "%Y-%m-%d")
        dt_fim = datetime.strptime(data_final.strip(), "%Y-%m-%d")
    except Exception:
        return "", ""

    if dt_fim < dt_ini:
        return "", ""

    diferenca_dias = (dt_fim.date() - dt_ini.date()).days + 1
    periodo_anterior_fim = dt_ini - timedelta(days=1)
    periodo_anterior_ini = periodo_anterior_fim - timedelta(days=diferenca_dias - 1)

    return (
        periodo_anterior_ini.strftime("%Y-%m-%d"),
        periodo_anterior_fim.strftime("%Y-%m-%d"),
    )


def calcular_variacao_percentual(valor_atual: int, valor_anterior: int) -> str:
    if valor_anterior == 0 and valor_atual == 0:
        return "0%"
    if valor_anterior == 0 and valor_atual > 0:
        return "+100%"
    variacao = ((valor_atual - valor_anterior) / valor_anterior) * 100
    sinal = "+" if variacao > 0 else ""
    return f"{sinal}{variacao:.0f}%"


def escape_html(valor):
    if valor is None:
        return ""
    return (
        str(valor)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#039;")
    )



def juntar_data_hora(data_cadastro: str, hora: str) -> str:
    try:
        data = data_cadastro.split(" ")[0].strip()
        return f"{data} {hora.strip()}"
    except Exception:
        return f"{data_cadastro.strip()} {hora.strip()}"


def _safe_json_dumps(data: Any) -> str:
    try:
        return json.dumps(data, ensure_ascii=False)
    except Exception:
        return str(data)


def _safe_json_loads(data: str) -> Dict[str, Any]:
    try:
        value = json.loads(data or "{}")
        return value if isinstance(value, dict) else {}
    except Exception:
        return {}


def _texto(payload: Dict[str, Any], *nomes: str, default: str = "") -> str:
    for nome in nomes:
        if "." in nome:
            grupo, chave = nome.split(".", 1)
            origem = payload.get(grupo) or {}
            valor = origem.get(chave) if isinstance(origem, dict) else None
        else:
            valor = payload.get(nome)
        if valor is not None and str(valor).strip() != "":
            return str(valor).strip()
    return default


def _booleano(payload: Dict[str, Any], *nomes: str, default: bool = False) -> bool:
    for nome in nomes:
        valor = payload.get(nome)
        if isinstance(valor, bool):
            return valor
        if isinstance(valor, int):
            return valor == 1
        if isinstance(valor, str):
            return valor.strip().lower() in ("1", "true", "sim", "s", "yes")
    return default


def _credenciais_payload(payload: Dict[str, Any]) -> Dict[str, str]:
    creds = payload.get("credenciais") or payload.get("credentials") or {}
    if not isinstance(creds, dict):
        creds = {}
    usuario = _texto(creds, "usuario")
    return {
        "empresa": _texto(creds, "empresa", default="1"),
        "usuario": usuario,
        "senha": _texto(creds, "senha"),
        "filial": _texto(creds, "filial", default=_texto(payload, "codigo_filial", "branch_code", default="1")),
        "recurso_humano": _texto(creds, "recurso_humano", "recurso", default=usuario),
    }


def _data_apenas(data_hora: str) -> str:
    data_hora = (data_hora or "").strip()
    return data_hora.split(" ")[0] if data_hora else ""


def montar_payload_os_frotaweb(payload: Dict[str, Any]) -> Dict[str, Any]:
    abertura = _texto(payload, "data_hora_abertura", "opening_datetime", "dh_entrada")
    saida = _texto(payload, "data_hora_saida", "exit_datetime", "dh_saida")
    hodometro = _texto(payload, "hodometro", "odometer", "km_entrada", default="0")

    return {
        "credenciais": _credenciais_payload(payload),
        "codigo_veiculo": _texto(payload, "codigo_veiculo", "vehicle_code", "cd_veiculo"),
        "descricao_defeito": _texto(payload, "descricao_defeito", "defect_description", "observacao", "observacoes"),
        "numero_os": _texto(payload, "numero_os", "order_number", default=""),
        "placa": _texto(payload, "placa", "plate").upper(),
        "codigo_componente": _texto(payload, "codigo_componente", "component_code", default=""),
        "data_abertura": _texto(payload, "data_abertura", default=_data_apenas(abertura)),
        "data_hora_abertura": abertura,
        "hodometro": hodometro,
        "horimetro_entrada": _texto(payload, "horimetro_entrada", "entry_hourmeter", default="0"),
        "data_hora_saida": saida,
        "hodometro_saida": _texto(payload, "hodometro_saida", "exit_odometer", default=hodometro),
        "horimetro_saida": _texto(payload, "horimetro_saida", "exit_hourmeter", default="0"),
        "data_hora_inicio": _texto(payload, "data_hora_inicio", "start_datetime", default=abertura),
        "data_hora_previsao_liberacao": _texto(payload, "data_hora_previsao_liberacao", "expected_release_datetime", default=saida),
        "horas_previstas": _texto(payload, "horas_previstas", "expected_hours", default="0.00"),
        "horas_realizadas": _texto(payload, "horas_realizadas", "actual_hours", "performed_hours", default="0"),
        "codigo_filial": _texto(payload, "codigo_filial", "branch_code", "cd_filial"),
        "codigo_departamento": _texto(payload, "codigo_departamento", "department_code", "cd_ccusto"),
        "codigo_oficina": _texto(payload, "codigo_oficina", "workshop_code", default=""),
        "codigo_servico": _texto(payload, "codigo_servico", "service_code", default=""),
        "codigo_solicitante": _texto(payload, "codigo_solicitante", "requester_code", default=""),
        "codigo_motorista": _texto(payload, "codigo_motorista", "driver_code", default="0"),
        "numero_ocorrencia": _texto(payload, "numero_ocorrencia", "occurrence_number", default="0"),
        "numero_contrato": _texto(payload, "numero_contrato", "contract_number", default=""),
        "valor_acrescimo": _texto(payload, "valor_acrescimo", "surcharge_value", "additional_value", default="0"),
        "numero_os_retorno": _texto(payload, "numero_os_retorno", "return_order_number", default="0"),
        "observacoes": _texto(payload, "observacoes", "observations", default=""),
        "investimento": _booleano(payload, "investimento", "investment"),
        "acidente": _booleano(payload, "acidente", "accident"),
        "socorro": _booleano(payload, "socorro", "roadside_assistance"),
        "servico_retorno": _booleano(payload, "servico_retorno", "return_service"),
        "programada": _booleano(payload, "programada", "scheduled"),
        "campos_brutos": payload.get("campos_brutos") if isinstance(payload.get("campos_brutos"), dict) else {},
    }


def montar_payload_servico_frotaweb(payload: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "credenciais": _credenciais_payload(payload),
        "numero_os": _texto(payload, "numero_os", "order_number"),
        "codigo_veiculo": _texto(payload, "codigo_veiculo", "vehicle_code", "cd_veiculo"),
        "placa": _texto(payload, "placa", "plate").upper(),
        "codigo_servico": _texto(payload, "codigo_servico", "service_code", default="0"),
        "codigo_recurso_humano": _texto(
            payload,
            "codigo_recurso_humano",
            "resource_code",
            "recurso_humano",
            default=_credenciais_payload(payload).get("recurso_humano", ""),
        ),
        "tempo_gasto": _texto(payload, "tempo_gasto", "spent_time", default="000:00"),
        "valor_hora": _texto(payload, "valor_hora", "hourly_value", default="0"),
        "campos_brutos": payload.get("campos_brutos") if isinstance(payload.get("campos_brutos"), dict) else {},
    }


def mascarar_senha(obj: Any) -> Any:
    if isinstance(obj, dict):
        novo = {}
        for chave, valor in obj.items():
            if str(chave).lower() == "senha":
                novo[chave] = "***"
            else:
                novo[chave] = mascarar_senha(valor)
        return novo
    if isinstance(obj, list):
        return [mascarar_senha(item) for item in obj]
    return obj


def post_json_frotaweb(url: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    try:
        resp = requests.post(url, json=payload, timeout=180)
    except requests.RequestException as exc:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar na API FrotaWeb: {str(exc)}")

    try:
        retorno = resp.json()
    except Exception:
        retorno = {"raw": resp.text[:3000]}

    if resp.status_code not in (200, 201):
        raise HTTPException(
            status_code=resp.status_code,
            detail={
                "url": url,
                "payload_enviado": mascarar_senha(payload),
                "retorno": retorno,
            },
        )

    return retorno


def extrair_numero_os(retorno: Any) -> str:
    candidatos = {
        "numero_os", "nr_os", "num_os", "os", "codigo_os", "cod_os",
        "order_number", "ordem_servico", "ordem", "id_os", "id",
    }

    def walk(obj):
        if isinstance(obj, dict):
            for chave, valor in obj.items():
                if str(chave).lower() in candidatos and valor not in (None, "", 0):
                    return str(valor).strip()
            for valor in obj.values():
                encontrado = walk(valor)
                if encontrado:
                    return encontrado
        if isinstance(obj, list):
            for item in obj:
                encontrado = walk(item)
                if encontrado:
                    return encontrado
        return ""

    return walk(retorno)


# =========================
# STARTUP
# =========================

@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)

    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS matricula VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS mecanico VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS km VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS hora_inicio VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS hora_fim VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS observacoes VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS data_cadastro VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS foto VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS frotaweb_status VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS frotaweb_numero_os VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS frotaweb_retorno VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico ADD COLUMN IF NOT EXISTS frotaweb_payload VARCHAR"))

        conn.execute(text("UPDATE ordens_servico SET frotaweb_status = 'PENDENTE' WHERE frotaweb_status IS NULL"))

        conn.execute(text("ALTER TABLE ordens_servico_itens ADD COLUMN IF NOT EXISTS frotaweb_status VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico_itens ADD COLUMN IF NOT EXISTS frotaweb_retorno VARCHAR"))
        conn.execute(text("ALTER TABLE ordens_servico_itens ADD COLUMN IF NOT EXISTS frotaweb_payload VARCHAR"))
        conn.execute(text("UPDATE ordens_servico_itens SET frotaweb_status = 'PENDENTE' WHERE frotaweb_status IS NULL"))

        conn.execute(text("UPDATE ordens_servico SET matricula = '' WHERE matricula IS NULL"))
        conn.execute(text("UPDATE ordens_servico SET mecanico = '' WHERE mecanico IS NULL"))
        conn.execute(text("UPDATE ordens_servico SET km = '' WHERE km IS NULL"))
        conn.execute(text("UPDATE ordens_servico SET hora_inicio = '' WHERE hora_inicio IS NULL"))
        conn.execute(text("UPDATE ordens_servico SET hora_fim = '' WHERE hora_fim IS NULL"))
        conn.execute(text("UPDATE ordens_servico SET observacoes = '' WHERE observacoes IS NULL"))
        conn.execute(text("UPDATE ordens_servico SET data_cadastro = '' WHERE data_cadastro IS NULL"))

        conn.execute(text("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS codigo VARCHAR"))
        conn.execute(text("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS referencia VARCHAR"))
        conn.execute(text("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS unidade VARCHAR"))
        conn.execute(text("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS ativo VARCHAR"))
        conn.execute(text("UPDATE produtos SET codigo = '' WHERE codigo IS NULL"))
        conn.execute(text("UPDATE produtos SET referencia = '' WHERE referencia IS NULL"))
        conn.execute(text("UPDATE produtos SET unidade = '' WHERE unidade IS NULL"))
        conn.execute(text("UPDATE produtos SET ativo = 'S' WHERE ativo IS NULL"))

        conn.execute(text("ALTER TABLE mecanicos ADD COLUMN IF NOT EXISTS codigo VARCHAR"))
        conn.execute(text("ALTER TABLE mecanicos ADD COLUMN IF NOT EXISTS nome VARCHAR"))
        conn.execute(text("ALTER TABLE mecanicos ADD COLUMN IF NOT EXISTS senha VARCHAR"))
        conn.execute(text("UPDATE mecanicos SET codigo = '' WHERE codigo IS NULL"))
        conn.execute(text("UPDATE mecanicos SET nome = '' WHERE nome IS NULL"))
        conn.execute(text("UPDATE mecanicos SET senha = '' WHERE senha IS NULL"))

        conn.execute(text("ALTER TABLE servicos ADD COLUMN IF NOT EXISTS codigo VARCHAR"))
        conn.execute(text("ALTER TABLE servicos ADD COLUMN IF NOT EXISTS sistema VARCHAR"))
        conn.execute(text("UPDATE servicos SET codigo = '' WHERE codigo IS NULL"))
        conn.execute(text("UPDATE servicos SET sistema = '' WHERE sistema IS NULL"))


        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS veiculos (
                id SERIAL PRIMARY KEY,
                codigo_frota VARCHAR NOT NULL,
                placa VARCHAR NOT NULL UNIQUE,
                cd_ccusto VARCHAR NOT NULL,
                cd_filial VARCHAR NOT NULL
            )
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS frotaweb_config (
                id SERIAL PRIMARY KEY,
                matricula VARCHAR NOT NULL UNIQUE,
                usuario VARCHAR NOT NULL,
                senha VARCHAR NOT NULL,
                filial_login VARCHAR NOT NULL,
                cd_empresa VARCHAR NOT NULL DEFAULT '1'
            )
        """))

    db = SessionLocal()
    try:
        servicos_iniciais = [
            {"codigo": "001", "nome": "Troca de óleo", "sistema": "Motor"},
            {"codigo": "002", "nome": "Troca de filtro", "sistema": "Motor"},
            {"codigo": "003", "nome": "Revisão geral", "sistema": "Geral"},
            {"codigo": "004", "nome": "Freios", "sistema": "Freio"},
            {"codigo": "005", "nome": "Suspensão", "sistema": "Suspensão"},
            {"codigo": "006", "nome": "Motor", "sistema": "Motor"},
            {"codigo": "007", "nome": "Elétrica", "sistema": "Elétrico"},
            {"codigo": "008", "nome": "Pneus", "sistema": "Rodagem"},
            {"codigo": "009", "nome": "Alinhamento", "sistema": "Suspensão"},
            {"codigo": "010", "nome": "Balanceamento", "sistema": "Rodagem"},
            {"codigo": "011", "nome": "Troca de embreagem", "sistema": "Transmissão"},
            {"codigo": "012", "nome": "Troca de bateria", "sistema": "Elétrico"},
        ]

        for item in servicos_iniciais:
            existe = db.query(ServicoModel).filter(ServicoModel.nome == item["nome"]).first()
            if not existe:
                db.add(ServicoModel(codigo=item["codigo"], nome=item["nome"], sistema=item["sistema"]))

        produtos_iniciais = [
            {"codigo": "1001", "nome": "Óleo 15W40", "referencia": "OLEO-15W40", "unidade": "L", "ativo": "S"},
            {"codigo": "1002", "nome": "Filtro de óleo", "referencia": "FILT-OL", "unidade": "UN", "ativo": "S"},
            {"codigo": "1003", "nome": "Filtro de ar", "referencia": "FILT-AR", "unidade": "UN", "ativo": "S"},
            {"codigo": "1004", "nome": "Pastilha de freio", "referencia": "PST-FR", "unidade": "JG", "ativo": "S"},
            {"codigo": "1005", "nome": "Graxa", "referencia": "GRX-01", "unidade": "KG", "ativo": "S"},
        ]

        for item in produtos_iniciais:
            existe = db.query(ProdutoModel).filter(ProdutoModel.nome == item["nome"]).first()
            if not existe:
                db.add(ProdutoModel(
                    codigo=item["codigo"], nome=item["nome"],
                    referencia=item["referencia"], unidade=item["unidade"], ativo=item["ativo"],
                ))

        mecanicos_iniciais = [
            {"codigo": "1001", "nome": "João Silva", "senha": "1234"},
            {"codigo": "1002", "nome": "Carlos Souza", "senha": "1234"},
            {"codigo": "1003", "nome": "Rafael Lima", "senha": "1234"},
            {"codigo": "1004", "nome": "Marcos Pereira", "senha": "1234"},
        ]

        for item in mecanicos_iniciais:
            existe = db.query(MecanicoModel).filter(MecanicoModel.codigo == item["codigo"]).first()
            if not existe:
                db.add(MecanicoModel(codigo=item["codigo"], nome=item["nome"], senha=item["senha"]))

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# =========================
# LOGIN
# =========================

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    if usuario_logado(request):
        return RedirectResponse(url="/painel/dashboard", status_code=303)
    return RedirectResponse(url="/login", status_code=303)




@app.get("/ui")
def ui_redirect():
    return RedirectResponse(url="/painel/dashboard", status_code=303)


@app.get("/ui/")
def ui_slash_redirect():
    return RedirectResponse(url="/painel/dashboard", status_code=303)

@app.get("/login", response_class=HTMLResponse)
def tela_login(request: Request, erro: str = ""):
    if usuario_logado(request):
        return RedirectResponse(url="/painel/dashboard", status_code=303)

    mensagem_erro = ""
    if erro:
        mensagem_erro = """
        <div style="background:#fee2e2;color:#991b1b;padding:12px;border-radius:8px;margin-bottom:12px;">
          Usuário ou senha inválidos
        </div>
        """

    html = f"""
    <html>
      <head>
        <meta charset="UTF-8">
        <title>Login do Painel</title>
        <style>
          body {{
            font-family: Arial, sans-serif;
            background: #f3f4f6;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
            margin: 0;
          }}
          .card {{
            background: white;
            padding: 28px;
            border-radius: 14px;
            box-shadow: 0 2px 16px rgba(0,0,0,0.10);
            width: 100%;
            max-width: 380px;
          }}
          h1 {{ margin-top: 0; margin-bottom: 20px; text-align: center; }}
          input {{
            width: 100%;
            padding: 12px;
            margin-bottom: 12px;
            border: 1px solid #d1d5db;
            border-radius: 8px;
            font-size: 14px;
            box-sizing: border-box;
          }}
          button {{
            width: 100%;
            padding: 12px;
            background: #1f2937;
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 15px;
            cursor: pointer;
          }}
        </style>
      </head>
      <body>
        <div class="card">
          <h1>Login do Painel</h1>
          {mensagem_erro}
          <form method="post" action="/login">
            <input type="text" name="usuario" placeholder="Usuário" required>
            <input type="password" name="senha" placeholder="Senha" required>
            <button type="submit">Entrar</button>
          </form>
        </div>
      </body>
    </html>
    """
    return html


@app.post("/login")
def login(request: Request, usuario: str = Form(...), senha: str = Form(...)):
    usuario_correto = "romulo@benel"
    senha_correta = "Caio2004@"

    if usuario.strip() == usuario_correto and senha.strip() == senha_correta:
        request.session["usuario_logado"] = True
        request.session["usuario_nome"] = usuario.strip()
        return RedirectResponse(url="/painel/dashboard", status_code=303)

    return RedirectResponse(url="/login?erro=1", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


# =========================
# ROTAS BÁSICAS
# =========================

@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/debug/tabelas")
def debug_tabelas(db: Session = Depends(get_db)):
    resultado = db.execute(
        text("""
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = 'public'
            ORDER BY table_name
        """)
    ).fetchall()
    return {"tables": [row[0] for row in resultado]}


# =========================
# SERVIÇOS
# =========================

@app.get("/servicos")
def listar_servicos(db: Session = Depends(get_db)):
    servicos = db.query(ServicoModel).order_by(ServicoModel.nome.asc()).all()
    return [
        {"id": item.id, "codigo": item.codigo or "", "nome": item.nome, "sistema": item.sistema or ""}
        for item in servicos
    ]


@app.post("/servicos", status_code=201)
def criar_servico(servico: ServicoCreate, db: Session = Depends(get_db)):
    nome = servico.nome.strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome do serviço é obrigatório")

    novo = ServicoModel(codigo=servico.codigo.strip(), nome=nome, sistema=servico.sistema.strip())
    db.add(novo)
    try:
        db.commit()
        db.refresh(novo)
        return {"id": novo.id, "codigo": novo.codigo or "", "nome": novo.nome, "sistema": novo.sistema or ""}
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Serviço já cadastrado")


@app.put("/servicos/{servico_id}")
def atualizar_servico(servico_id: int, servico: ServicoUpdate, db: Session = Depends(get_db)):
    registro = db.query(ServicoModel).filter(ServicoModel.id == servico_id).first()
    if registro is None:
        raise HTTPException(status_code=404, detail="Serviço não encontrado")

    registro.codigo = servico.codigo.strip()
    registro.nome = servico.nome.strip()
    registro.sistema = servico.sistema.strip()

    try:
        db.commit()
        db.refresh(registro)
        return {"id": registro.id, "codigo": registro.codigo or "", "nome": registro.nome, "sistema": registro.sistema or ""}
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Serviço já cadastrado")


@app.post("/servicos/{servico_id}/excluir")
def excluir_servico(servico_id: int, db: Session = Depends(get_db)):
    registro = db.query(ServicoModel).filter(ServicoModel.id == servico_id).first()
    if registro is None:
        raise HTTPException(status_code=404, detail="Serviço não encontrado")
    db.delete(registro)
    db.commit()
    return RedirectResponse(url="/painel/servicos", status_code=303)


@app.post("/servicos/importar-xlsx")
async def importar_servicos_xlsx(arquivo: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Pacote openpyxl não instalado no servidor")

    if not arquivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    conteudo = await arquivo.read()
    wb = load_workbook(filename=BytesIO(conteudo), data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        descricao = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        sistema = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not descricao:
            continue
        existente = db.query(ServicoModel).filter(ServicoModel.nome == descricao).first()
        if existente:
            existente.codigo = codigo
            existente.sistema = sistema
        else:
            db.add(ServicoModel(codigo=codigo, nome=descricao, sistema=sistema))

    db.commit()
    return RedirectResponse(url="/painel/servicos", status_code=303)


@app.post("/painel/servicos/adicionar")
def adicionar_servico_painel(
    request: Request,
    codigo: str = Form(""),
    nome: str = Form(...),
    sistema: str = Form(""),
    db: Session = Depends(get_db),
):
    if not usuario_logado(request):
        return RedirectResponse(url="/login", status_code=303)

    nome = nome.strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Descrição do serviço é obrigatória")

    novo = ServicoModel(codigo=codigo.strip(), nome=nome, sistema=sistema.strip())
    db.add(novo)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Serviço já cadastrado")

    return RedirectResponse(url="/painel/servicos", status_code=303)


@app.get("/painel/servicos", response_class=HTMLResponse)
def painel_servicos(request: Request, db: Session = Depends(get_db)):
    if not usuario_logado(request):
        return RedirectResponse(url="/login", status_code=303)

    servicos = db.query(ServicoModel).order_by(ServicoModel.nome.asc()).all()

    html = """
    <html>
      <head>
        <meta charset="UTF-8">
        <title>Painel de Serviços</title>
        <style>
          body { font-family: Arial, sans-serif; margin: 20px; background: #f7f7f7; }
          h1 { margin-bottom: 20px; }
          .card { background: #fff; border-radius: 12px; padding: 16px; box-shadow: 0 2px 10px rgba(0,0,0,0.08); margin-bottom: 20px; overflow-x: auto; }
          .acoes { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 16px; }
          .btn { padding: 10px 16px; border: none; border-radius: 8px; cursor: pointer; font-size: 14px; text-decoration: none; display: inline-block; }
          .btn-voltar { background: #e5e7eb; color: #111827; }
          .btn-salvar { background: #0f766e; color: white; }
          .btn-excluir { background: #dc2626; color: white; }
          .btn-importar { background: #1d4ed8; color: white; }
          .form-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 12px; margin-bottom: 12px; }
          .form-grid input { padding: 10px; border: 1px solid #ccc; border-radius: 8px; font-size: 14px; }
          table { width: 100%; border-collapse: collapse; }
          th, td { border: 1px solid #ddd; padding: 10px; text-align: left; }
          th { background: #1f2937; color: white; }
          tr:nth-child(even) { background: #fafafa; }
          form.inline { display: inline; }
          .upload-box { display: flex; gap: 10px; flex-wrap: wrap; align-items: center; margin-top: 12px; }
          input[type=file] { padding: 8px; border: 1px solid #ccc; border-radius: 8px; background: white; }
        </style>
      </head>
      <body>
        <h1>Painel de Serviços</h1>
        <div class="card">
          <div class="acoes">
            <a href="/painel/dashboard" class="btn btn-voltar">Ir para dashboard</a>
            <a href="/painel/ordens-servico" class="btn btn-voltar">Ir para ordens de serviço</a>
            <a href="/painel/produtos" class="btn btn-voltar">Ir para produtos</a>
            <a href="/painel/mecanicos" class="btn btn-voltar">Ir para mecânicos</a>
            <a href="/logout" class="btn btn-voltar">Sair</a>
          </div>
          <form method="post" action="/painel/servicos/adicionar">
            <div class="form-grid">
              <input type="text" name="codigo" placeholder="Código do serviço">
              <input type="text" name="nome" placeholder="Descrição do serviço" required>
              <input type="text" name="sistema" placeholder="Sistema do serviço">
            </div>
            <button type="submit" class="btn btn-salvar">Adicionar serviço</button>
          </form>
          <form method="post" action="/servicos/importar-xlsx" enctype="multipart/form-data">
            <div class="upload-box">
              <input type="file" name="arquivo" accept=".xlsx" required>
              <button type="submit" class="btn btn-importar">Importar serviços XLSX</button>
            </div>
            <p style="margin-top:10px; color:#555;">
              Estrutura esperada: coluna A = código, coluna B = descrição, coluna C = sistema.
            </p>
          </form>
        </div>
        <div class="card">
          <table>
            <thead>
              <tr><th>ID</th><th>Código</th><th>Descrição</th><th>Sistema</th><th>Ações</th></tr>
            </thead>
            <tbody>
    """

    if servicos:
        for item in servicos:
            html += f"""
              <tr>
                <td>{item.id}</td>
                <td>{escape_html(item.codigo or '')}</td>
                <td>{escape_html(item.nome)}</td>
                <td>{escape_html(item.sistema or '')}</td>
                <td>
                  <form class="inline" method="post" action="/servicos/{item.id}/excluir"
                        onsubmit="return confirm('Deseja excluir {escape_html(item.nome)}?');">
                    <button type="submit" class="btn btn-excluir">Excluir</button>
                  </form>
                </td>
              </tr>
            """
    else:
        html += "<tr><td colspan='5'>Nenhum serviço cadastrado</td></tr>"

    html += "</tbody></table></div></body></html>"
    return html


# =========================
# MECÂNICOS
# =========================

@app.get("/mecanicos")
def listar_mecanicos(db: Session = Depends(get_db)):
    mecanicos = db.query(MecanicoModel).order_by(MecanicoModel.nome.asc()).all()
    return [{"id": item.id, "codigo": item.codigo, "nome": item.nome, "senha": item.senha} for item in mecanicos]


@app.post("/mecanicos", status_code=201)
def criar_mecanico(mecanico: MecanicoCreate, db: Session = Depends(get_db)):
    codigo = mecanico.codigo.strip()
    nome = mecanico.nome.strip()
    senha = mecanico.senha.strip()

    if not codigo:
        raise HTTPException(status_code=400, detail="Código é obrigatório")
    if not nome:
        raise HTTPException(status_code=400, detail="Nome é obrigatório")
    if not senha:
        raise HTTPException(status_code=400, detail="Senha é obrigatória")

    novo = MecanicoModel(codigo=codigo, nome=nome, senha=senha)
    db.add(novo)
    try:
        db.commit()
        db.refresh(novo)
        return {"id": novo.id, "codigo": novo.codigo, "nome": novo.nome, "senha": novo.senha}
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Código do mecânico já cadastrado")


@app.put("/mecanicos/{mecanico_id}")
def atualizar_mecanico(mecanico_id: int, mecanico: MecanicoUpdate, db: Session = Depends(get_db)):
    registro = db.query(MecanicoModel).filter(MecanicoModel.id == mecanico_id).first()
    if registro is None:
        raise HTTPException(status_code=404, detail="Mecânico não encontrado")

    registro.codigo = mecanico.codigo.strip()
    registro.nome = mecanico.nome.strip()
    registro.senha = mecanico.senha.strip()

    try:
        db.commit()
        db.refresh(registro)
        return {"id": registro.id, "codigo": registro.codigo, "nome": registro.nome, "senha": registro.senha}
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Código do mecânico já cadastrado")


@app.post("/mecanicos/{mecanico_id}/excluir")
def excluir_mecanico(mecanico_id: int, db: Session = Depends(get_db)):
    registro = db.query(MecanicoModel).filter(MecanicoModel.id == mecanico_id).first()
    if registro is None:
        raise HTTPException(status_code=404, detail="Mecânico não encontrado")
    db.delete(registro)
    db.commit()
    return RedirectResponse(url="/painel/mecanicos", status_code=303)


@app.get("/painel/mecanicos", response_class=HTMLResponse)
def painel_mecanicos(request: Request, db: Session = Depends(get_db)):
    if not usuario_logado(request):
        return RedirectResponse(url="/login", status_code=303)

    mecanicos = db.query(MecanicoModel).order_by(MecanicoModel.nome.asc()).all()

    html = """
    <html>
      <head>
        <meta charset="UTF-8">
        <title>Painel de Mecânicos</title>
        <style>
          body { font-family: Arial, sans-serif; margin: 20px; background: #f7f7f7; }
          h1 { margin-bottom: 20px; }
          .card { background: #fff; border-radius: 12px; padding: 16px; box-shadow: 0 2px 10px rgba(0,0,0,0.08); margin-bottom: 20px; overflow-x: auto; }
          .acoes { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 16px; }
          .btn { padding: 10px 16px; border: none; border-radius: 8px; cursor: pointer; font-size: 14px; text-decoration: none; display: inline-block; }
          .btn-voltar { background: #e5e7eb; color: #111827; }
          .btn-salvar { background: #0f766e; color: white; }
          .btn-excluir { background: #dc2626; color: white; }
          .form-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 12px; margin-bottom: 12px; }
          .form-grid input { padding: 10px; border: 1px solid #ccc; border-radius: 8px; font-size: 14px; }
          table { width: 100%; border-collapse: collapse; }
          th, td { border: 1px solid #ddd; padding: 10px; text-align: left; }
          th { background: #1f2937; color: white; }
          tr:nth-child(even) { background: #fafafa; }
          form.inline { display: inline; }
        </style>
      </head>
      <body>
        <h1>Painel de Mecânicos</h1>
        <div class="card">
          <div class="acoes">
            <a href="/painel/dashboard" class="btn btn-voltar">Ir para dashboard</a>
            <a href="/painel/ordens-servico" class="btn btn-voltar">Ir para ordens de serviço</a>
            <a href="/painel/produtos" class="btn btn-voltar">Ir para produtos</a>
            <a href="/painel/servicos" class="btn btn-voltar">Ir para serviços</a>
            <a href="/logout" class="btn btn-voltar">Sair</a>
          </div>
          <form method="post" action="/painel/mecanicos/adicionar">
            <div class="form-grid">
              <input type="text" name="codigo" placeholder="Código de identificação" required>
              <input type="text" name="nome" placeholder="Nome completo" required>
              <input type="text" name="senha" placeholder="Senha" required>
            </div>
            <button type="submit" class="btn btn-salvar">Adicionar mecânico</button>
          </form>
        </div>
        <div class="card">
          <table>
            <thead>
              <tr><th>ID</th><th>Código</th><th>Nome</th><th>Senha</th><th>Ações</th></tr>
            </thead>
            <tbody>
    """

    if mecanicos:
        for item in mecanicos:
            html += f"""
              <tr>
                <td>{item.id}</td>
                <td>{escape_html(item.codigo)}</td>
                <td>{escape_html(item.nome)}</td>
                <td>{escape_html(item.senha)}</td>
                <td>
                  <form class="inline" method="post" action="/mecanicos/{item.id}/excluir"
                        onsubmit="return confirm('Deseja excluir {escape_html(item.nome)}?');">
                    <button type="submit" class="btn btn-excluir">Excluir</button>
                  </form>
                </td>
              </tr>
            """
    else:
        html += "<tr><td colspan='5'>Nenhum mecânico cadastrado</td></tr>"

    html += "</tbody></table></div></body></html>"
    return html


@app.post("/painel/mecanicos/adicionar")
def adicionar_mecanico_painel(
    request: Request,
    codigo: str = Form(...),
    nome: str = Form(...),
    senha: str = Form(...),
    db: Session = Depends(get_db),
):
    if not usuario_logado(request):
        return RedirectResponse(url="/login", status_code=303)

    codigo = codigo.strip()
    nome = nome.strip()
    senha = senha.strip()

    if not codigo or not nome or not senha:
        raise HTTPException(status_code=400, detail="Preencha código, nome e senha")

    novo = MecanicoModel(codigo=codigo, nome=nome, senha=senha)
    db.add(novo)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Código do mecânico já cadastrado")

    return RedirectResponse(url="/painel/mecanicos", status_code=303)


# =========================
# PRODUTOS
# =========================

@app.get("/produtos")
def listar_produtos(db: Session = Depends(get_db)):
    produtos = (
        db.query(ProdutoModel)
        .filter(ProdutoModel.ativo == "S")
        .order_by(ProdutoModel.nome.asc())
        .all()
    )
    return [
        {
            "id": item.id, "codigo": item.codigo or "", "nome": item.nome,
            "referencia": item.referencia or "", "unidade": item.unidade or "", "ativo": item.ativo or "S",
        }
        for item in produtos
    ]


@app.post("/produtos", status_code=201)
def criar_produto(produto: ProdutoCreate, db: Session = Depends(get_db)):
    nome = produto.nome.strip()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome do produto é obrigatório")

    novo = ProdutoModel(
        codigo=produto.codigo.strip(), nome=nome,
        referencia=produto.referencia.strip(), unidade=produto.unidade.strip(),
        ativo=(produto.ativo or "S").strip() or "S",
    )
    db.add(novo)
    try:
        db.commit()
        db.refresh(novo)
        return {
            "id": novo.id, "codigo": novo.codigo or "", "nome": novo.nome,
            "referencia": novo.referencia or "", "unidade": novo.unidade or "", "ativo": novo.ativo or "S",
        }
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Produto já cadastrado")


@app.put("/produtos/{produto_id}")
def atualizar_produto(produto_id: int, produto: ProdutoUpdate, db: Session = Depends(get_db)):
    registro = db.query(ProdutoModel).filter(ProdutoModel.id == produto_id).first()
    if registro is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    registro.codigo = produto.codigo.strip()
    registro.nome = produto.nome.strip()
    registro.referencia = produto.referencia.strip()
    registro.unidade = produto.unidade.strip()
    registro.ativo = (produto.ativo or "S").strip() or "S"

    try:
        db.commit()
        db.refresh(registro)
        return {
            "id": registro.id, "codigo": registro.codigo or "", "nome": registro.nome,
            "referencia": registro.referencia or "", "unidade": registro.unidade or "", "ativo": registro.ativo or "S",
        }
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Produto já cadastrado")


@app.post("/produtos/{produto_id}/excluir")
def excluir_produto(produto_id: int, db: Session = Depends(get_db)):
    registro = db.query(ProdutoModel).filter(ProdutoModel.id == produto_id).first()
    if registro is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    db.delete(registro)
    db.commit()
    return RedirectResponse(url="/painel/produtos", status_code=303)


@app.post("/produtos/importar-xlsx")
async def importar_produtos_xlsx(arquivo: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Pacote openpyxl não instalado no servidor")

    if not arquivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    conteudo = await arquivo.read()
    wb = load_workbook(filename=BytesIO(conteudo), data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        descricao = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        referencia = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not descricao:
            continue
        existente = db.query(ProdutoModel).filter(ProdutoModel.nome == descricao).first()
        if existente:
            existente.codigo = codigo
            existente.referencia = referencia
            existente.ativo = "S"
        else:
            db.add(ProdutoModel(codigo=codigo, nome=descricao, referencia=referencia, unidade="", ativo="S"))

    db.commit()
    return RedirectResponse(url="/painel/produtos", status_code=303)


@app.get("/painel/produtos", response_class=HTMLResponse)
def painel_produtos(request: Request, db: Session = Depends(get_db)):
    if not usuario_logado(request):
        return RedirectResponse(url="/login", status_code=303)

    produtos = db.query(ProdutoModel).order_by(ProdutoModel.nome.asc()).all()

    html = """
    <html>
      <head>
        <meta charset="UTF-8">
        <title>Painel de Produtos</title>
        <style>
          body { font-family: Arial, sans-serif; margin: 20px; background: #f7f7f7; }
          h1 { margin-bottom: 20px; }
          .card { background: #fff; border-radius: 12px; padding: 16px; box-shadow: 0 2px 10px rgba(0,0,0,0.08); margin-bottom: 20px; overflow-x: auto; }
          table { width: 100%; border-collapse: collapse; }
          th, td { border: 1px solid #ddd; padding: 10px; text-align: left; }
          th { background: #1f2937; color: white; }
          tr:nth-child(even) { background: #fafafa; }
          .acoes { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 16px; }
          .btn { padding: 10px 16px; border: none; border-radius: 8px; cursor: pointer; font-size: 14px; text-decoration: none; display: inline-block; }
          .btn-voltar { background: #e5e7eb; color: #111827; }
          .btn-excluir { background: #dc2626; color: white; }
          .btn-importar { background: #0f766e; color: white; }
          form.inline { display: inline; }
          .upload-box { display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }
          input[type=file] { padding: 8px; border: 1px solid #ccc; border-radius: 8px; background: white; }
        </style>
      </head>
      <body>
        <h1>Painel de Produtos</h1>
        <div class="card">
          <div class="acoes">
            <a href="/painel/dashboard" class="btn btn-voltar">Ir para dashboard</a>
            <a href="/painel/ordens-servico" class="btn btn-voltar">Ir para ordens de serviço</a>
            <a href="/painel/mecanicos" class="btn btn-voltar">Ir para mecânicos</a>
            <a href="/painel/servicos" class="btn btn-voltar">Ir para serviços</a>
            <a href="/logout" class="btn btn-voltar">Sair</a>
          </div>
          <form method="post" action="/produtos/importar-xlsx" enctype="multipart/form-data">
            <div class="upload-box">
              <input type="file" name="arquivo" accept=".xlsx" required>
              <button type="submit" class="btn btn-importar">Importar base XLSX</button>
            </div>
            <p style="margin-top:10px; color:#555;">
              Estrutura esperada: coluna A = código, coluna B = descrição, coluna C = referência.
            </p>
          </form>
        </div>
        <div class="card">
          <table>
            <thead>
              <tr><th>ID</th><th>Código</th><th>Descrição</th><th>Referência</th><th>Unidade</th><th>Ativo</th><th>Ações</th></tr>
            </thead>
            <tbody>
    """

    if produtos:
        for item in produtos:
            html += f"""
              <tr>
                <td>{item.id}</td>
                <td>{escape_html(item.codigo or '')}</td>
                <td>{escape_html(item.nome)}</td>
                <td>{escape_html(item.referencia or '')}</td>
                <td>{escape_html(item.unidade or '')}</td>
                <td>{escape_html(item.ativo or 'S')}</td>
                <td>
                  <form class="inline" method="post" action="/produtos/{item.id}/excluir"
                        onsubmit="return confirm('Deseja excluir {escape_html(item.nome)}?');">
                    <button type="submit" class="btn btn-excluir">Excluir</button>
                  </form>
                </td>
              </tr>
            """
    else:
        html += "<tr><td colspan='7'>Nenhum produto cadastrado</td></tr>"

    html += "</tbody></table></div></body></html>"
    return html


# =========================
# VEÍCULOS
# =========================

@app.get("/veiculos")
def listar_veiculos(db: Session = Depends(get_db)):
    veiculos = db.query(VeiculoModel).order_by(VeiculoModel.placa.asc()).all()
    return [
        {
            "id": item.id,
            "codigo_frota": item.codigo_frota,
            "placa": item.placa,
            "cd_ccusto": item.cd_ccusto,
            "cd_filial": item.cd_filial,
        }
        for item in veiculos
    ]


@app.post("/veiculos", status_code=201)
def criar_veiculo(veiculo: VeiculoCreate, db: Session = Depends(get_db)):
    novo = VeiculoModel(
        codigo_frota=veiculo.codigo_frota.strip(),
        placa=veiculo.placa.strip().upper(),
        cd_ccusto=veiculo.cd_ccusto.strip(),
        cd_filial=veiculo.cd_filial.strip(),
    )

    db.add(novo)
    try:
        db.commit()
        db.refresh(novo)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Veículo já cadastrado")

    return {
        "id": novo.id,
        "codigo_frota": novo.codigo_frota,
        "placa": novo.placa,
        "cd_ccusto": novo.cd_ccusto,
        "cd_filial": novo.cd_filial,
    }


@app.put("/veiculos/{veiculo_id}")
def atualizar_veiculo(veiculo_id: int, veiculo: VeiculoUpdate, db: Session = Depends(get_db)):
    registro = db.query(VeiculoModel).filter(VeiculoModel.id == veiculo_id).first()
    if registro is None:
        raise HTTPException(status_code=404, detail="Veículo não encontrado")

    registro.codigo_frota = veiculo.codigo_frota.strip()
    registro.placa = veiculo.placa.strip().upper()
    registro.cd_ccusto = veiculo.cd_ccusto.strip()
    registro.cd_filial = veiculo.cd_filial.strip()

    try:
        db.commit()
        db.refresh(registro)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Placa já cadastrada")

    return {
        "id": registro.id,
        "codigo_frota": registro.codigo_frota,
        "placa": registro.placa,
        "cd_ccusto": registro.cd_ccusto,
        "cd_filial": registro.cd_filial,
    }


# =========================
# CONFIG FROTAWEB
# =========================

@app.get("/frotaweb-config/{matricula}")
def obter_frotaweb_config(matricula: str, db: Session = Depends(get_db)):
    config = db.query(FrotaWebConfigModel).filter(
        FrotaWebConfigModel.matricula == matricula.strip()
    ).first()

    if config is None:
        raise HTTPException(status_code=404, detail="Configuração FrotaWeb não encontrada para esta matrícula")

    return {
        "id": config.id,
        "matricula": config.matricula,
        "usuario": config.usuario,
        "senha": config.senha,
        "filial_login": config.filial_login,
        "cd_empresa": config.cd_empresa,
    }


@app.post("/frotaweb-config", status_code=201)
def criar_frotaweb_config(config: FrotaWebConfigCreate, db: Session = Depends(get_db)):
    novo = FrotaWebConfigModel(
        matricula=config.matricula.strip(),
        usuario=config.usuario.strip(),
        senha=config.senha.strip(),
        filial_login=config.filial_login.strip(),
        cd_empresa=config.cd_empresa.strip() or "1",
    )

    db.add(novo)
    try:
        db.commit()
        db.refresh(novo)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Já existe configuração FrotaWeb para essa matrícula")

    return {
        "id": novo.id,
        "matricula": novo.matricula,
        "usuario": novo.usuario,
        "senha": novo.senha,
        "filial_login": novo.filial_login,
        "cd_empresa": novo.cd_empresa,
    }


@app.put("/frotaweb-config/{matricula}")
def atualizar_frotaweb_config(matricula: str, config: FrotaWebConfigUpdate, db: Session = Depends(get_db)):
    registro = db.query(FrotaWebConfigModel).filter(
        FrotaWebConfigModel.matricula == matricula.strip()
    ).first()

    if registro is None:
        raise HTTPException(status_code=404, detail="Configuração FrotaWeb não encontrada")

    registro.usuario = config.usuario.strip()
    registro.senha = config.senha.strip()
    registro.filial_login = config.filial_login.strip()
    registro.cd_empresa = config.cd_empresa.strip() or "1"

    db.commit()
    db.refresh(registro)

    return {
        "id": registro.id,
        "matricula": registro.matricula,
        "usuario": registro.usuario,
        "senha": registro.senha,
        "filial_login": registro.filial_login,
        "cd_empresa": registro.cd_empresa,
    }


# =========================
# INTEGRAÇÃO FROTAWEB
# =========================

INTEGRACAO_FROTAWEB_URL = "http://127.0.0.1:8000/frotaweb/os-corretiva"


@app.post("/integracoes/frotaweb/os-corretiva")
def integrar_os_frotaweb(payload: OrdemServicoFrotaWebRequest, db: Session = Depends(get_db)):
    config = db.query(FrotaWebConfigModel).filter(
        FrotaWebConfigModel.matricula == payload.matricula.strip()
    ).first()

    if config is None:
        raise HTTPException(
            status_code=404,
            detail="Não existe configuração FrotaWeb cadastrada para esta matrícula"
        )

    veiculo = db.query(VeiculoModel).filter(
        VeiculoModel.id == payload.veiculo_id
    ).first()

    if veiculo is None:
        raise HTTPException(status_code=404, detail="Veículo não encontrado")

    if not payload.cd_servicos:
        raise HTTPException(status_code=400, detail="Informe pelo menos um código de serviço")

    dh_entrada = juntar_data_hora(payload.data_cadastro, payload.hora_inicio)
    dh_saida = juntar_data_hora(payload.data_cadastro, payload.hora_fim)
    dh_inicio = juntar_data_hora(payload.data_cadastro, payload.hora_inicio)
    dh_prev = juntar_data_hora(payload.data_cadastro, payload.hora_fim)

    body_integracao = {
        "usuario": config.usuario,
        "senha": config.senha,
        "filial_login": config.filial_login,
        "cd_empresa": config.cd_empresa,
        "cd_veiculo": veiculo.codigo_frota,
        "placa": veiculo.placa,
        "dh_entrada": dh_entrada,
        "km_entrada": payload.km.strip(),
        "dh_saida": dh_saida,
        "km_saida": payload.km.strip(),
        "dh_inicio": dh_inicio,
        "dh_prev": dh_prev,
        "cd_filial": veiculo.cd_filial,
        "cd_ccusto": veiculo.cd_ccusto,
        "observacao": payload.observacoes.strip(),
        "cd_servicos": payload.cd_servicos,
    }

    try:
        resp = requests.post(
            INTEGRACAO_FROTAWEB_URL,
            json=body_integracao,
            timeout=120,
        )
    except requests.RequestException as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Falha ao conectar na API de integração FrotaWeb: {str(exc)}"
        )

    try:
        retorno = resp.json()
    except Exception:
        retorno = {"raw": resp.text}

    if resp.status_code not in (200, 201):
        raise HTTPException(
            status_code=resp.status_code,
            detail=retorno,
        )

    return {
        "ok": True,
        "payload_enviado": body_integracao,
        "retorno_integracao": retorno,
    }


# =========================
# ORDENS DE SERVIÇO
# =========================

@app.post("/ordens-servico", status_code=201)
async def criar_ordem_servico(
    placa: str = Form(...),
    matricula: str = Form(...),
    mecanico: str = Form(...),
    km: str = Form(...),
    hora_inicio: str = Form(...),
    hora_fim: str = Form(...),
    servicos: str = Form(...),
    produtos: str = Form("[]"),
    observacoes: str = Form(""),
    data_cadastro: str = Form(...),
    foto: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    import json

    lista_servicos = [s.strip() for s in json.loads(servicos) if s.strip()]
    if not lista_servicos:
        raise HTTPException(status_code=400, detail="Informe pelo menos um serviço")

    lista_produtos = json.loads(produtos)

    foto_path = None
    if foto and foto.filename:
        ext = os.path.splitext(foto.filename)[-1].lower() or ".jpg"
        nome_arquivo = f"{int(datetime.now().timestamp() * 1000)}{ext}"
        destino = os.path.join(UPLOAD_DIR, nome_arquivo)
        with open(destino, "wb") as f:
            shutil.copyfileobj(foto.file, f)
        foto_path = f"/fotos/{nome_arquivo}"

    nova_ordem = OrdemServicoModel(
        placa=placa.strip(),
        matricula=matricula.strip(),
        mecanico=mecanico.strip(),
        km=km.strip(),
        hora_inicio=hora_inicio.strip(),
        hora_fim=hora_fim.strip(),
        observacoes=observacoes.strip(),
        data_cadastro=data_cadastro.strip(),
        foto=foto_path,
    )

    db.add(nova_ordem)
    db.flush()

    for nome_servico in lista_servicos:
        db.add(OrdemServicoItemModel(ordem_id=nova_ordem.id, servico=nome_servico))

    for p in lista_produtos:
        nome_p = (p.get("produto_nome") or "").strip()
        qtd = (p.get("quantidade") or "").strip()
        if nome_p and qtd:
            db.add(OrdemServicoProdutoModel(
                ordem_id=nova_ordem.id,
                produto_id=p.get("produto_id"),
                produto_nome=nome_p,
                quantidade=qtd,
            ))

    db.commit()
    db.refresh(nova_ordem)

    return {
        "message": "Ordem criada com sucesso",
        "id": nova_ordem.id,
        "foto_url": foto_path,
        "total_itens": len(lista_servicos),
        "total_produtos": len([p for p in lista_produtos if p.get("produto_nome") and p.get("quantidade")]),
    }


@app.post("/ordens-servico/{ordem_id}/excluir")
def excluir_ordem_servico(ordem_id: int, db: Session = Depends(get_db)):
    ordem = db.query(OrdemServicoModel).filter(OrdemServicoModel.id == ordem_id).first()
    if ordem is None:
        raise HTTPException(status_code=404, detail="Ordem de serviço não encontrada")

    # Remove o arquivo de foto do disco se existir
    if ordem.foto:
        caminho_fisico = ordem.foto.lstrip("/").replace("fotos/", f"{UPLOAD_DIR}/", 1)
        if os.path.exists(caminho_fisico):
            os.remove(caminho_fisico)

    db.delete(ordem)
    db.commit()
    return RedirectResponse(url="/painel/ordens-servico", status_code=303)


@app.get("/ordens-servico")
def listar_ordens_servico(
    placa: str = "",
    matricula: str = "",
    mecanico: str = "",
    km: str = "",
    servico: str = "",
    data_cadastro: str = "",
    db: Session = Depends(get_db),
):
    try:
        rows = buscar_linhas_ordens(db, placa=placa, matricula=matricula, mecanico=mecanico, km=km, servico=servico, data_cadastro=data_cadastro)
        return [
            {
                "id": row.id, "placa": row.placa, "matricula": row.matricula,
                "mecanico": row.mecanico, "km": row.km, "hora_inicio": row.hora_inicio,
                "hora_fim": row.hora_fim, "servico": row.servico,
                "observacoes": row.observacoes, "data_cadastro": row.data_cadastro,
                "foto_url": row.foto,
                "frotaweb_status": getattr(row, "frotaweb_status", "PENDENTE"),
                "frotaweb_numero_os": getattr(row, "frotaweb_numero_os", ""),
                "frotaweb_retorno": getattr(row, "frotaweb_retorno", ""),
            }
            for row in rows
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao listar ordens: {str(e)}")



# =========================
# ROTAS DO APP PARA O PAINEL
# =========================
# Estas rotas recebem os dados enviados pelo app Android e salvam somente no painel.
# Elas NAO enviam automaticamente para o FrotaWeb.
# Fluxo correto: App Android -> /panel/os -> Painel /ui/ -> conferencia -> envio manual ao FrotaWeb.


def _app_text(payload: Dict[str, Any], *names: str, default: str = "") -> str:
    for name in names:
        value = payload.get(name)
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return default


def _app_bool(payload: Dict[str, Any], *names: str) -> bool:
    for name in names:
        value = payload.get(name)
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().lower() in ("1", "true", "sim", "s", "yes")
        if isinstance(value, int):
            return value == 1
    return False


def _app_credentials(payload: Dict[str, Any]) -> Dict[str, Any]:
    creds = payload.get("credenciais") or payload.get("credentials") or {}
    return creds if isinstance(creds, dict) else {}


def _normalizar_data_painel(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return datetime.now().strftime("%d/%m/%Y %H:%M")

    formatos = (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    )
    for fmt in formatos:
        try:
            return datetime.strptime(value, fmt).strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass

    # Se nao conseguiu interpretar, remove segundos quando existir.
    partes = value.split()
    if len(partes) >= 2 and partes[1].count(":") >= 2:
        return f"{partes[0]} {'/'.join([])}".replace(" ", " ") if False else f"{partes[0]} {':'.join(partes[1].split(':')[:2])}"
    return value


def _hora_painel(value: str) -> str:
    value = _normalizar_data_painel(value)
    partes = value.split()
    if len(partes) >= 2:
        hora = partes[1]
        pedacos = hora.split(":")
        if len(pedacos) >= 2:
            return f"{pedacos[0].zfill(2)}:{pedacos[1].zfill(2)}"
    return "00:00"


def _salvar_primeira_foto_app(payload: Dict[str, Any]) -> Optional[str]:
    photos = payload.get("photos") or payload.get("fotos") or []
    if not isinstance(photos, list) or not photos:
        return None

    first = photos[0]
    if not isinstance(first, dict):
        return None

    data_base64 = first.get("data_base64") or first.get("base64") or first.get("data")
    if not data_base64:
        return None

    try:
        raw = base64.b64decode(str(data_base64), validate=False)
    except Exception:
        return None

    filename = str(first.get("filename") or first.get("nome") or "foto-os.jpg")
    ext = os.path.splitext(filename)[-1].lower()
    if ext not in (".jpg", ".jpeg", ".png"):
        ext = ".jpg"

    nome_arquivo = f"app_{int(datetime.now().timestamp() * 1000)}{ext}"
    destino = os.path.join(UPLOAD_DIR, nome_arquivo)
    with open(destino, "wb") as f:
        f.write(raw)
    return f"/fotos/{nome_arquivo}"


@app.post("/panel/os")
async def receber_os_app_no_painel(payload: Dict[str, Any], db: Session = Depends(get_db)):
    """
    Recebe uma O.S. do app Android e salva no painel para conferencia.
    Nao envia ao FrotaWeb.
    """
    creds = _app_credentials(payload)

    placa = _app_text(payload, "plate", "placa")
    codigo_veiculo = _app_text(payload, "vehicle_code", "codigo_veiculo", "cd_veiculo")
    descricao_defeito = _app_text(payload, "defect_description", "descricao_defeito")
    abertura = _app_text(payload, "opening_datetime", "data_hora_abertura", "data_abertura")
    saida = _app_text(payload, "exit_datetime", "data_hora_saida")
    km = _app_text(payload, "odometer", "hodometro", "km", default="0")
    observacoes = _app_text(payload, "observations", "observacoes")

    matricula = _app_text(
        payload,
        "matricula",
        "resource_code",
        "recurso_humano",
        default=str(creds.get("recurso_humano") or creds.get("recurso") or creds.get("usuario") or "").strip(),
    )
    mecanico = _app_text(
        payload,
        "mecanico",
        "mechanic",
        default=str(creds.get("usuario") or matricula or "").strip(),
    )

    if not placa:
        raise HTTPException(status_code=400, detail="Placa obrigatoria para salvar a O.S. no painel")

    detalhes = []
    if codigo_veiculo:
        detalhes.append(f"Veiculo/Frota: {codigo_veiculo}")
    if descricao_defeito:
        detalhes.append(f"Defeito: {descricao_defeito}")
    if _app_text(payload, "branch_code", "codigo_filial"):
        detalhes.append(f"Filial O.S.: {_app_text(payload, 'branch_code', 'codigo_filial')}")
    if _app_text(payload, "department_code", "codigo_departamento"):
        detalhes.append(f"Departamento: {_app_text(payload, 'department_code', 'codigo_departamento')}")
    if _app_bool(payload, "investment", "investimento"):
        detalhes.append("Investimento: Sim")
    if _app_bool(payload, "accident", "acidente"):
        detalhes.append("Acidente: Sim")
    if _app_bool(payload, "roadside_assistance", "socorro"):
        detalhes.append("Socorro: Sim")
    if _app_bool(payload, "return_service", "servico_retorno"):
        detalhes.append("Retorno: Sim")
    if _app_bool(payload, "scheduled", "programada"):
        detalhes.append("Programada: Sim")

    observacao_final = observacoes.strip()
    if detalhes:
        observacao_final = (observacao_final + " | " if observacao_final else "") + " | ".join(detalhes)

    foto_path = _salvar_primeira_foto_app(payload)

    payload_frotaweb = montar_payload_os_frotaweb(payload)

    nova_ordem = OrdemServicoModel(
        placa=placa.strip().upper(),
        matricula=matricula.strip(),
        mecanico=mecanico.strip(),
        km=km.strip(),
        hora_inicio=_hora_painel(abertura),
        hora_fim=_hora_painel(saida),
        observacoes=observacao_final,
        data_cadastro=_normalizar_data_painel(abertura),
        foto=foto_path,
        frotaweb_status="PENDENTE",
        frotaweb_payload=_safe_json_dumps(payload_frotaweb),
    )

    db.add(nova_ordem)
    db.flush()

    servico_base = _app_text(
        payload,
        "service_code",
        "codigo_servico",
        default="Corretiva pendente de conferencia",
    )
    db.add(OrdemServicoItemModel(
        ordem_id=nova_ordem.id,
        servico=servico_base,
        frotaweb_status="PENDENTE",
    ))

    db.commit()
    db.refresh(nova_ordem)

    return {
        "created": True,
        "accepted": False,
        "order_number": str(nova_ordem.id),
        "panel_id": str(nova_ordem.id),
        "message": f"O.S. {nova_ordem.id} salva no painel para conferencia. Nao foi enviada ao FrotaWeb.",
    }


@app.post("/panel/os/servicos")
async def receber_servico_app_no_painel(payload: Dict[str, Any], db: Session = Depends(get_db)):
    """
    Recebe servico realizado no app e vincula na O.S. do painel.
    Nao envia ao FrotaWeb.
    """
    numero_os = _app_text(payload, "order_number", "numero_os")
    codigo_servico = _app_text(payload, "service_code", "codigo_servico", default="Servico realizado")
    tempo = _app_text(payload, "spent_time", "horas_realizadas")

    if not numero_os:
        raise HTTPException(status_code=400, detail="Numero da O.S. obrigatorio para salvar o servico no painel")

    try:
        ordem_id = int(str(numero_os).replace("Painel #", "").strip())
    except Exception:
        raise HTTPException(status_code=400, detail="Numero da O.S. invalido para o painel")

    ordem = db.query(OrdemServicoModel).filter(OrdemServicoModel.id == ordem_id).first()
    if ordem is None:
        raise HTTPException(status_code=404, detail="O.S. nao encontrada no painel")

    descricao = codigo_servico.strip() or "Servico realizado"
    if tempo:
        descricao = f"{descricao} - Tempo: {tempo}"

    payload_servico = montar_payload_servico_frotaweb(payload)
    numero_os_frotaweb = ordem.frotaweb_numero_os or str(ordem.id)
    payload_servico["numero_os"] = payload_servico.get("numero_os") or numero_os_frotaweb
    if not payload_servico.get("codigo_veiculo"):
        payload_os = _safe_json_loads(ordem.frotaweb_payload or "{}")
        payload_servico["codigo_veiculo"] = payload_os.get("codigo_veiculo", "")
    if not payload_servico.get("placa"):
        payload_servico["placa"] = ordem.placa or ""

    db.add(OrdemServicoItemModel(
        ordem_id=ordem.id,
        servico=descricao,
        frotaweb_status="PENDENTE",
        frotaweb_payload=_safe_json_dumps(payload_servico),
    ))
    db.commit()

    return {
        "created": True,
        "accepted": False,
        "order_number": str(ordem.id),
        "message": f"Servico salvo no painel na O.S. {ordem.id}. Nao foi enviado ao FrotaWeb.",
    }


@app.post("/painel/ordens-servico/{ordem_id}/enviar-frotaweb")
def enviar_ordem_painel_para_frotaweb(request: Request, ordem_id: int, db: Session = Depends(get_db)):
    """
    Envia manualmente uma O.S. do painel para a API FrotaWeb.
    Depois de enviar a O.S., tenta enviar automaticamente os serviços vinculados.
    """
    if not usuario_logado(request):
        return RedirectResponse(url="/login", status_code=303)

    ordem = db.query(OrdemServicoModel).filter(OrdemServicoModel.id == ordem_id).first()
    if ordem is None:
        raise HTTPException(status_code=404, detail="O.S. não encontrada")

    payload_os = _safe_json_loads(ordem.frotaweb_payload or "{}")
    if not payload_os:
        payload_os = montar_payload_os_frotaweb({
            "placa": ordem.placa,
            "hodometro": ordem.km,
            "data_hora_abertura": juntar_data_hora(ordem.data_cadastro or "", ordem.hora_inicio or ""),
            "data_hora_saida": juntar_data_hora(ordem.data_cadastro or "", ordem.hora_fim or ""),
            "observacoes": ordem.observacoes or "",
            "credenciais": {
                "empresa": "1",
                "usuario": ordem.mecanico or ordem.matricula or "",
                "senha": "",
                "filial": "",
                "recurso_humano": ordem.matricula or "",
            }
        })

    try:
        retorno_os = post_json_frotaweb(FROTAWEB_OS_URL, payload_os)
    except HTTPException as exc:
        ordem.frotaweb_status = "ERRO_ENVIO"
        ordem.frotaweb_retorno = _safe_json_dumps(exc.detail)
        db.commit()
        return RedirectResponse(url="/painel/ordens-servico?servico=Erro%20FrotaWeb", status_code=303)

    numero_os = extrair_numero_os(retorno_os)
    if numero_os:
        ordem.frotaweb_numero_os = numero_os

    servicos = db.query(OrdemServicoItemModel).filter(OrdemServicoItemModel.ordem_id == ordem.id).all()
    total_servicos = len(servicos)
    servicos_enviados = 0
    servicos_erros = 0

    for item in servicos:
        # Não envia item base quando não houver payload real de serviço salvo pelo app.
        payload_servico = _safe_json_loads(item.frotaweb_payload or "{}")
        if not payload_servico:
            item.frotaweb_status = item.frotaweb_status or "PENDENTE"
            continue

        payload_servico["numero_os"] = numero_os or ordem.frotaweb_numero_os or str(ordem.id)

        try:
            retorno_servico = post_json_frotaweb(FROTAWEB_SERVICO_URL, payload_servico)
            item.frotaweb_status = "ENVIADA"
            item.frotaweb_retorno = _safe_json_dumps(mascarar_senha(retorno_servico))
            servicos_enviados += 1
        except HTTPException as exc:
            item.frotaweb_status = "ERRO_ENVIO"
            item.frotaweb_retorno = _safe_json_dumps(exc.detail)
            servicos_erros += 1

    ordem.frotaweb_status = "ENVIADA"
    ordem.frotaweb_retorno = (
        f"O.S. enviada com sucesso. "
        f"Nº O.S. FrotaWeb: {ordem.frotaweb_numero_os or 'não identificado'}. "
        f"Serviços: {servicos_enviados}/{total_servicos} enviados, {servicos_erros} erro(s). "
        f"Retorno: {_safe_json_dumps(mascarar_senha(retorno_os))}"
    )
    db.commit()

    return RedirectResponse(url="/painel/ordens-servico", status_code=303)


@app.post("/ordens-servico/{ordem_id}/enviar-frotaweb")
def enviar_ordem_painel_para_frotaweb_json(ordem_id: int, db: Session = Depends(get_db)):
    """
    Versão JSON do envio manual para testes via /docs.
    """
    ordem = db.query(OrdemServicoModel).filter(OrdemServicoModel.id == ordem_id).first()
    if ordem is None:
        raise HTTPException(status_code=404, detail="O.S. não encontrada")

    payload_os = _safe_json_loads(ordem.frotaweb_payload or "{}")
    if not payload_os:
        raise HTTPException(status_code=400, detail="O.S. não possui payload FrotaWeb salvo")

    retorno_os = post_json_frotaweb(FROTAWEB_OS_URL, payload_os)
    numero_os = extrair_numero_os(retorno_os)
    if numero_os:
        ordem.frotaweb_numero_os = numero_os

    servicos = db.query(OrdemServicoItemModel).filter(OrdemServicoItemModel.ordem_id == ordem.id).all()
    resumo_servicos = []
    for item in servicos:
        payload_servico = _safe_json_loads(item.frotaweb_payload or "{}")
        if not payload_servico:
            resumo_servicos.append({"id": item.id, "status": "SEM_PAYLOAD"})
            continue
        payload_servico["numero_os"] = numero_os or ordem.frotaweb_numero_os or str(ordem.id)
        try:
            retorno_servico = post_json_frotaweb(FROTAWEB_SERVICO_URL, payload_servico)
            item.frotaweb_status = "ENVIADA"
            item.frotaweb_retorno = _safe_json_dumps(mascarar_senha(retorno_servico))
            resumo_servicos.append({"id": item.id, "status": "ENVIADA", "retorno": mascarar_senha(retorno_servico)})
        except HTTPException as exc:
            item.frotaweb_status = "ERRO_ENVIO"
            item.frotaweb_retorno = _safe_json_dumps(exc.detail)
            resumo_servicos.append({"id": item.id, "status": "ERRO_ENVIO", "erro": exc.detail})

    ordem.frotaweb_status = "ENVIADA"
    ordem.frotaweb_retorno = _safe_json_dumps(mascarar_senha(retorno_os))
    db.commit()

    return {
        "ok": True,
        "numero_os_frotaweb": ordem.frotaweb_numero_os,
        "retorno_os": mascarar_senha(retorno_os),
        "servicos": resumo_servicos,
    }


# =========================
# DASHBOARD JSON
# =========================

@app.get("/dashboard/dados")
def dashboard_dados(data_inicial: str = "", data_final: str = "", db: Session = Depends(get_db)):
    ordens = buscar_ordens_dashboard(db, data_inicial=data_inicial, data_final=data_final)

    total_os = len(ordens)
    total_servicos = 0
    total_produtos = 0
    placas_unicas = set()
    soma_tempo = 0

    servicos_contagem = defaultdict(int)
    mecanicos_contagem = defaultdict(int)
    veiculos_contagem = defaultdict(int)
    veiculos_tempo = defaultdict(int)
    servicos_tempo = defaultdict(int)
    servicos_qtd_tempo = defaultdict(int)
    produtos_contagem = defaultdict(float)

    for ordem in ordens:
        placa = ordem.placa or "Não informado"
        mecanico = ordem.mecanico or "Não informado"

        placas_unicas.add(placa)
        mecanicos_contagem[mecanico] += 1
        veiculos_contagem[placa] += 1

        duracao = calcular_duracao_minutos(ordem.hora_inicio or "", ordem.hora_fim or "")
        soma_tempo += duracao
        veiculos_tempo[placa] += duracao

        for item in ordem.itens:
            nome_servico = item.servico or "Não informado"
            total_servicos += 1
            servicos_contagem[nome_servico] += 1
            servicos_tempo[nome_servico] += duracao
            servicos_qtd_tempo[nome_servico] += 1

        for produto in ordem.produtos_consumidos:
            total_produtos += 1
            try:
                qtd = float(str(produto.quantidade).replace(",", "."))
            except Exception:
                qtd = 0.0
            produtos_contagem[produto.produto_nome or "Não informado"] += qtd

    tempo_medio = int(soma_tempo / total_os) if total_os > 0 else 0
    media_servicos_por_os = round(total_servicos / total_os, 2) if total_os > 0 else 0
    media_produtos_por_os = round(total_produtos / total_os, 2) if total_os > 0 else 0

    servicos_mais_realizados = sorted(
        [{"servico": k, "quantidade": v} for k, v in servicos_contagem.items()],
        key=lambda x: x["quantidade"], reverse=True,
    )[:10]

    mecanicos_produtivos = sorted(
        [{"mecanico": k, "quantidade": v} for k, v in mecanicos_contagem.items()],
        key=lambda x: x["quantidade"], reverse=True,
    )[:10]

    veiculos_mais_entradas = sorted(
        [{"placa": k, "quantidade": v} for k, v in veiculos_contagem.items()],
        key=lambda x: x["quantidade"], reverse=True,
    )[:10]

    veiculos_maior_tempo = sorted(
        [{"placa": k, "minutos": v, "tempo_formatado": minutos_para_texto(v)} for k, v in veiculos_tempo.items()],
        key=lambda x: x["minutos"], reverse=True,
    )[:10]

    tempo_medio_por_servico = []
    for nome_servico, minutos_total in servicos_tempo.items():
        qtd = servicos_qtd_tempo[nome_servico]
        media = int(minutos_total / qtd) if qtd > 0 else 0
        tempo_medio_por_servico.append({
            "servico": nome_servico, "media_minutos": media, "media_formatada": minutos_para_texto(media),
        })
    tempo_medio_por_servico = sorted(tempo_medio_por_servico, key=lambda x: x["media_minutos"], reverse=True)[:10]

    produtos_mais_consumidos = sorted(
        [{"produto": k, "quantidade": v} for k, v in produtos_contagem.items()],
        key=lambda x: x["quantidade"], reverse=True,
    )[:10]

    maior_mecanico = mecanicos_produtivos[0] if mecanicos_produtivos else {"mecanico": "-", "quantidade": 0}
    maior_veiculo_entradas = veiculos_mais_entradas[0] if veiculos_mais_entradas else {"placa": "-", "quantidade": 0}
    maior_veiculo_tempo = veiculos_maior_tempo[0] if veiculos_maior_tempo else {"placa": "-", "tempo_formatado": "0h 00min", "minutos": 0}
    maior_servico = servicos_mais_realizados[0] if servicos_mais_realizados else {"servico": "-", "quantidade": 0}
    maior_produto = produtos_mais_consumidos[0] if produtos_mais_consumidos else {"produto": "-", "quantidade": 0}

    return {
        "resumo": {
            "total_os": total_os, "total_servicos": total_servicos, "total_produtos": total_produtos,
            "veiculos_unicos": len([p for p in placas_unicas if p]),
            "tempo_medio_minutos": tempo_medio, "tempo_medio_formatado": minutos_para_texto(tempo_medio),
            "media_servicos_por_os": media_servicos_por_os, "media_produtos_por_os": media_produtos_por_os,
        },
        "destaques": {
            "mecanico_top": maior_mecanico, "veiculo_top_entradas": maior_veiculo_entradas,
            "veiculo_top_tempo": maior_veiculo_tempo, "servico_top": maior_servico, "produto_top": maior_produto,
        },
        "servicos_mais_realizados": servicos_mais_realizados,
        "mecanicos_produtivos": mecanicos_produtivos,
        "veiculos_mais_entradas": veiculos_mais_entradas,
        "veiculos_maior_tempo": veiculos_maior_tempo,
        "tempo_medio_por_servico": tempo_medio_por_servico,
        "produtos_mais_consumidos": produtos_mais_consumidos,
    }


# =========================
# DASHBOARD HTML
# =========================

@app.get("/painel/dashboard", response_class=HTMLResponse)
def painel_dashboard(
    request: Request,
    data_inicial: str = "",
    data_final: str = "",
    db: Session = Depends(get_db),
):
    if not usuario_logado(request):
        return RedirectResponse(url="/login", status_code=303)

    dados = dashboard_dados(data_inicial=data_inicial, data_final=data_final, db=db)

    resumo = dados["resumo"]
    destaques = dados["destaques"]
    servicos_mais_realizados = dados["servicos_mais_realizados"]
    mecanicos_produtivos = dados["mecanicos_produtivos"]
    veiculos_mais_entradas = dados["veiculos_mais_entradas"]
    veiculos_maior_tempo = dados["veiculos_maior_tempo"]
    tempo_medio_por_servico = dados["tempo_medio_por_servico"]
    produtos_mais_consumidos = dados["produtos_mais_consumidos"]

    data_inicial_anterior, data_final_anterior = obter_intervalo_periodo_anterior(data_inicial, data_final)
    dados_periodo_anterior = None
    if data_inicial_anterior and data_final_anterior:
        dados_periodo_anterior = dashboard_dados(data_inicial=data_inicial_anterior, data_final=data_final_anterior, db=db)

    comparativo_total_os = ""
    comparativo_total_servicos = ""
    comparativo_total_produtos = ""
    comparativo_tempo_medio = ""

    if dados_periodo_anterior:
        resumo_anterior = dados_periodo_anterior["resumo"]
        comparativo_total_os = calcular_variacao_percentual(resumo["total_os"], resumo_anterior["total_os"])
        comparativo_total_servicos = calcular_variacao_percentual(resumo["total_servicos"], resumo_anterior["total_servicos"])
        comparativo_total_produtos = calcular_variacao_percentual(resumo["total_produtos"], resumo_anterior["total_produtos"])
        comparativo_tempo_medio = calcular_variacao_percentual(resumo["tempo_medio_minutos"], resumo_anterior["tempo_medio_minutos"])

    periodo_exibicao = "Todo o período"
    if data_inicial and data_final:
        periodo_exibicao = f"{data_inicial} até {data_final}"
    elif data_inicial:
        periodo_exibicao = f"A partir de {data_inicial}"
    elif data_final:
        periodo_exibicao = f"Até {data_final}"

    usuario_nome = request.session.get("usuario_nome", "Usuário")

    html = f"""
    <html>
      <head>
        <meta charset="UTF-8">
        <title>Dashboard da Oficina</title>
        <style>
          * {{ box-sizing: border-box; }}
          body {{ font-family: Arial, sans-serif; margin: 0; background: #eef2f7; color: #111827; }}
          .container {{ max-width: 1600px; margin: 0 auto; padding: 24px; }}
          .hero {{ background: linear-gradient(135deg, #111827 0%, #1f2937 100%); color: white; border-radius: 18px; padding: 24px; box-shadow: 0 10px 24px rgba(15,23,42,0.18); margin-bottom: 20px; }}
          .hero-top {{ display: flex; justify-content: space-between; align-items: flex-start; gap: 16px; flex-wrap: wrap; }}
          .hero h1 {{ margin: 0 0 8px 0; font-size: 32px; }}
          .hero p {{ margin: 0; color: #d1d5db; font-size: 14px; }}
          .hero-badge {{ background: rgba(255,255,255,0.10); border: 1px solid rgba(255,255,255,0.12); border-radius: 12px; padding: 12px 16px; min-width: 240px; }}
          .hero-badge strong {{ display: block; font-size: 14px; color: #f9fafb; margin-bottom: 4px; }}
          .hero-badge span {{ font-size: 13px; color: #d1d5db; }}
          .card {{ background: #ffffff; border-radius: 16px; padding: 18px; box-shadow: 0 6px 18px rgba(15,23,42,0.06); margin-bottom: 20px; }}
          .section-title {{ margin: 0 0 16px 0; font-size: 18px; font-weight: 700; }}
          .section-subtitle {{ margin: -8px 0 16px 0; font-size: 13px; color: #6b7280; }}
          .filtros-form {{ display: flex; flex-direction: column; gap: 16px; }}
          .filtros-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 14px; }}
          .filtros-grid label {{ display: flex; flex-direction: column; gap: 8px; font-size: 13px; color: #374151; font-weight: 600; }}
          .filtros-grid input {{ padding: 12px 14px; border: 1px solid #d1d5db; border-radius: 10px; font-size: 14px; background: #fff; }}
          .acoes {{ display: flex; gap: 10px; flex-wrap: wrap; }}
          .btn {{ padding: 11px 16px; border: none; border-radius: 10px; cursor: pointer; font-size: 14px; text-decoration: none; display: inline-block; font-weight: 600; }}
          .btn-primario {{ background: #111827; color: white; }}
          .btn-secundario {{ background: #e5e7eb; color: #111827; }}
          .btn-sucesso {{ background: #0f766e; color: white; }}
          .btn-azul {{ background: #1d4ed8; color: white; }}
          .grid-kpis {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 16px; margin-bottom: 20px; }}
          .kpi {{ background: #fff; border-radius: 16px; padding: 18px; box-shadow: 0 6px 18px rgba(15,23,42,0.06); border: 1px solid #edf2f7; }}
          .kpi-label {{ font-size: 13px; color: #6b7280; margin-bottom: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.4px; }}
          .kpi-value {{ font-size: 30px; font-weight: 800; color: #111827; line-height: 1.1; margin-bottom: 8px; }}
          .kpi-foot {{ font-size: 13px; color: #6b7280; }}
          .kpi-variation {{ display: inline-block; margin-top: 8px; padding: 5px 10px; border-radius: 999px; font-size: 12px; font-weight: 700; background: #f3f4f6; color: #111827; }}
          .grid-insights {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 16px; margin-bottom: 20px; }}
          .insight {{ background: #fff; border-radius: 16px; padding: 18px; box-shadow: 0 6px 18px rgba(15,23,42,0.06); border-left: 6px solid #111827; }}
          .insight h3 {{ margin: 0 0 10px 0; font-size: 14px; color: #374151; }}
          .insight .principal {{ font-size: 20px; font-weight: 800; color: #111827; margin-bottom: 6px; }}
          .insight .secundario {{ font-size: 13px; color: #6b7280; }}
          .blocos-2 {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(440px, 1fr)); gap: 20px; margin-bottom: 20px; }}
          .blocos-3 {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 20px; margin-bottom: 20px; }}
          .table-wrap {{ overflow-x: auto; }}
          table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
          th, td {{ border-bottom: 1px solid #e5e7eb; padding: 12px 10px; text-align: left; vertical-align: middle; }}
          th {{ background: #111827; color: white; font-size: 13px; position: sticky; top: 0; }}
          tr:nth-child(even) {{ background: #fafafa; }}
          tr:hover {{ background: #f3f8ff; }}
          .td-num {{ text-align: right; font-weight: 700; }}
          .empty {{ padding: 20px 10px; text-align: center; color: #6b7280; }}
          .footer-note {{ color: #6b7280; font-size: 12px; margin-top: 10px; }}
          @media (max-width: 768px) {{ .container {{ padding: 14px; }} .hero h1 {{ font-size: 24px; }} .kpi-value {{ font-size: 24px; }} }}
        </style>
        <script>
          function setPeriodo(tipo) {{
            const hoje = new Date();
            const campoInicial = document.getElementById("data_inicial");
            const campoFinal = document.getElementById("data_final");
            function formatar(dt) {{
              return `${{dt.getFullYear()}}-${{String(dt.getMonth()+1).padStart(2,'0')}}-${{String(dt.getDate()).padStart(2,'0')}}`;
            }}
            if (tipo === "hoje") {{ campoInicial.value = formatar(hoje); campoFinal.value = formatar(hoje); }}
            if (tipo === "7dias") {{ const i = new Date(hoje); i.setDate(hoje.getDate()-6); campoInicial.value = formatar(i); campoFinal.value = formatar(hoje); }}
            if (tipo === "30dias") {{ const i = new Date(hoje); i.setDate(hoje.getDate()-29); campoInicial.value = formatar(i); campoFinal.value = formatar(hoje); }}
            if (tipo === "mes") {{ campoInicial.value = formatar(new Date(hoje.getFullYear(), hoje.getMonth(), 1)); campoFinal.value = formatar(new Date(hoje.getFullYear(), hoje.getMonth()+1, 0)); }}
          }}
        </script>
      </head>
      <body>
        <div class="container">
          <div class="hero">
            <div class="hero-top">
              <div>
                <h1>Dashboard da Oficina</h1>
                <p>Visão gerencial da operação de ordens de serviço, consumo e produtividade.</p>
              </div>
              <div class="hero-badge">
                <strong>Período aplicado</strong>
                <span>{escape_html(periodo_exibicao)}</span>
                <strong style="margin-top:10px;">Usuário logado</strong>
                <span>{escape_html(usuario_nome)}</span>
              </div>
            </div>
          </div>

          <div class="card">
            <h2 class="section-title">Filtros e navegação</h2>
            <p class="section-subtitle">Use os filtros para recalcular todos os indicadores do painel.</p>
            <form method="get" action="/painel/dashboard" class="filtros-form">
              <div class="filtros-grid">
                <label>Data inicial<input type="date" id="data_inicial" name="data_inicial" value="{escape_html(data_inicial)}"></label>
                <label>Data final<input type="date" id="data_final" name="data_final" value="{escape_html(data_final)}"></label>
              </div>
              <div class="acoes">
                <button type="submit" class="btn btn-primario">Aplicar filtro</button>
                <a href="/painel/dashboard" class="btn btn-secundario">Limpar filtro</a>
                <button type="button" class="btn btn-secundario" onclick="setPeriodo('hoje')">Hoje</button>
                <button type="button" class="btn btn-secundario" onclick="setPeriodo('7dias')">Últimos 7 dias</button>
                <button type="button" class="btn btn-secundario" onclick="setPeriodo('30dias')">Últimos 30 dias</button>
                <button type="button" class="btn btn-secundario" onclick="setPeriodo('mes')">Mês atual</button>
              </div>
              <div class="acoes">
                <a href="/painel/ordens-servico" class="btn btn-azul">Ir para painel de O.S.</a>
                <a href="/painel/produtos" class="btn btn-secundario">Ir para produtos</a>
                <a href="/painel/mecanicos" class="btn btn-secundario">Ir para mecânicos</a>
                <a href="/painel/servicos" class="btn btn-secundario">Ir para serviços</a>
                <a href="/logout" class="btn btn-secundario">Sair</a>
              </div>
            </form>
          </div>

          <div class="grid-kpis">
            <div class="kpi">
              <div class="kpi-label">Total de O.S.</div>
              <div class="kpi-value">{resumo["total_os"]}</div>
              <div class="kpi-foot">Ordens registradas no período filtrado</div>
              {f'<div class="kpi-variation">vs período anterior: {comparativo_total_os}</div>' if comparativo_total_os else ''}
            </div>
            <div class="kpi">
              <div class="kpi-label">Serviços executados</div>
              <div class="kpi-value">{resumo["total_servicos"]}</div>
              <div class="kpi-foot">Soma de todos os serviços lançados</div>
              {f'<div class="kpi-variation">vs período anterior: {comparativo_total_servicos}</div>' if comparativo_total_servicos else ''}
            </div>
            <div class="kpi">
              <div class="kpi-label">Produtos lançados</div>
              <div class="kpi-value">{resumo["total_produtos"]}</div>
              <div class="kpi-foot">Itens consumidos vinculados às O.S.</div>
              {f'<div class="kpi-variation">vs período anterior: {comparativo_total_produtos}</div>' if comparativo_total_produtos else ''}
            </div>
            <div class="kpi">
              <div class="kpi-label">Veículos únicos</div>
              <div class="kpi-value">{resumo["veiculos_unicos"]}</div>
              <div class="kpi-foot">Placas distintas atendidas</div>
            </div>
            <div class="kpi">
              <div class="kpi-label">Tempo médio por O.S.</div>
              <div class="kpi-value">{escape_html(resumo["tempo_medio_formatado"])}</div>
              <div class="kpi-foot">Calculado por hora de início e fim</div>
              {f'<div class="kpi-variation">vs período anterior: {comparativo_tempo_medio}</div>' if comparativo_tempo_medio else ''}
            </div>
            <div class="kpi">
              <div class="kpi-label">Média de serviços por O.S.</div>
              <div class="kpi-value">{resumo["media_servicos_por_os"]}</div>
              <div class="kpi-foot">Quantidade média de serviços em cada ordem</div>
            </div>
            <div class="kpi">
              <div class="kpi-label">Média de produtos por O.S.</div>
              <div class="kpi-value">{resumo["media_produtos_por_os"]}</div>
              <div class="kpi-foot">Quantidade média de produtos por ordem</div>
            </div>
          </div>

          <div class="grid-insights">
            <div class="insight"><h3>Serviço mais realizado</h3><div class="principal">{escape_html(destaques["servico_top"]["servico"])}</div><div class="secundario">{destaques["servico_top"]["quantidade"]} lançamento(s)</div></div>
            <div class="insight"><h3>Mecânico com mais O.S.</h3><div class="principal">{escape_html(destaques["mecanico_top"]["mecanico"])}</div><div class="secundario">{destaques["mecanico_top"]["quantidade"]} ordem(ns)</div></div>
            <div class="insight"><h3>Veículo com mais entradas</h3><div class="principal">{escape_html(destaques["veiculo_top_entradas"]["placa"])}</div><div class="secundario">{destaques["veiculo_top_entradas"]["quantidade"]} entrada(s)</div></div>
            <div class="insight"><h3>Veículo com maior tempo parado</h3><div class="principal">{escape_html(destaques["veiculo_top_tempo"]["placa"])}</div><div class="secundario">{escape_html(destaques["veiculo_top_tempo"]["tempo_formatado"])}</div></div>
            <div class="insight"><h3>Produto mais consumido</h3><div class="principal">{escape_html(destaques["produto_top"]["produto"])}</div><div class="secundario">Quantidade: {destaques["produto_top"]["quantidade"]}</div></div>
          </div>

          <div class="blocos-2">
            <div class="card">
              <h2 class="section-title">Serviços mais realizados</h2>
              <p class="section-subtitle">Ranking dos serviços com maior recorrência.</p>
              <div class="table-wrap"><table><thead><tr><th>Serviço</th><th style="text-align:right;">Quantidade</th></tr></thead><tbody>
    """

    if servicos_mais_realizados:
        for item in servicos_mais_realizados:
            html += f"<tr><td>{escape_html(item['servico'])}</td><td class='td-num'>{item['quantidade']}</td></tr>"
    else:
        html += "<tr><td colspan='2' class='empty'>Nenhum dado encontrado</td></tr>"

    html += """</tbody></table></div></div>
            <div class="card">
              <h2 class="section-title">Mecânicos com mais O.S.</h2>
              <p class="section-subtitle">Produtividade por quantidade de ordens registradas.</p>
              <div class="table-wrap"><table><thead><tr><th>Mecânico</th><th style="text-align:right;">Quantidade</th></tr></thead><tbody>
    """

    if mecanicos_produtivos:
        for item in mecanicos_produtivos:
            html += f"<tr><td>{escape_html(item['mecanico'])}</td><td class='td-num'>{item['quantidade']}</td></tr>"
    else:
        html += "<tr><td colspan='2' class='empty'>Nenhum dado encontrado</td></tr>"

    html += """</tbody></table></div></div></div>
          <div class="blocos-3">
            <div class="card">
              <h2 class="section-title">Produtos mais consumidos</h2>
              <p class="section-subtitle">Volume acumulado por produto.</p>
              <div class="table-wrap"><table><thead><tr><th>Produto</th><th style="text-align:right;">Quantidade</th></tr></thead><tbody>
    """

    if produtos_mais_consumidos:
        for item in produtos_mais_consumidos:
            html += f"<tr><td>{escape_html(item['produto'])}</td><td class='td-num'>{item['quantidade']}</td></tr>"
    else:
        html += "<tr><td colspan='2' class='empty'>Nenhum dado encontrado</td></tr>"

    html += """</tbody></table></div></div>
            <div class="card">
              <h2 class="section-title">Tempo médio por serviço</h2>
              <p class="section-subtitle">Serviços mais demorados no período.</p>
              <div class="table-wrap"><table><thead><tr><th>Serviço</th><th style="text-align:right;">Tempo médio</th></tr></thead><tbody>
    """

    if tempo_medio_por_servico:
        for item in tempo_medio_por_servico:
            html += f"<tr><td>{escape_html(item['servico'])}</td><td class='td-num'>{escape_html(item['media_formatada'])}</td></tr>"
    else:
        html += "<tr><td colspan='2' class='empty'>Nenhum dado encontrado</td></tr>"

    html += f"""</tbody></table></div></div>
            <div class="card">
              <h2 class="section-title">Resumo operacional</h2>
              <p class="section-subtitle">Leitura rápida da oficina.</p>
              <div style="display:grid; gap:12px;">
                <div style="padding:14px; background:#f9fafb; border:1px solid #e5e7eb; border-radius:12px;">
                  <div style="font-size:12px; color:#6b7280; font-weight:700; text-transform:uppercase;">Serviços por O.S.</div>
                  <div style="font-size:24px; font-weight:800; margin-top:6px;">{resumo["media_servicos_por_os"]}</div>
                </div>
                <div style="padding:14px; background:#f9fafb; border:1px solid #e5e7eb; border-radius:12px;">
                  <div style="font-size:12px; color:#6b7280; font-weight:700; text-transform:uppercase;">Produtos por O.S.</div>
                  <div style="font-size:24px; font-weight:800; margin-top:6px;">{resumo["media_produtos_por_os"]}</div>
                </div>
                <div style="padding:14px; background:#f9fafb; border:1px solid #e5e7eb; border-radius:12px;">
                  <div style="font-size:12px; color:#6b7280; font-weight:700; text-transform:uppercase;">Tempo médio por O.S.</div>
                  <div style="font-size:24px; font-weight:800; margin-top:6px;">{escape_html(resumo["tempo_medio_formatado"])}</div>
                </div>
              </div>
            </div>
          </div>

          <div class="blocos-2">
            <div class="card">
              <h2 class="section-title">Veículos com mais entradas</h2>
              <p class="section-subtitle">Placas com maior reincidência de atendimento.</p>
              <div class="table-wrap"><table><thead><tr><th>Placa</th><th style="text-align:right;">Quantidade</th></tr></thead><tbody>
    """

    if veiculos_mais_entradas:
        for item in veiculos_mais_entradas:
            html += f"<tr><td>{escape_html(item['placa'])}</td><td class='td-num'>{item['quantidade']}</td></tr>"
    else:
        html += "<tr><td colspan='2' class='empty'>Nenhum dado encontrado</td></tr>"

    html += """</tbody></table></div></div>
            <div class="card">
              <h2 class="section-title">Veículos com maior tempo parado</h2>
              <p class="section-subtitle">Tempo acumulado por placa.</p>
              <div class="table-wrap"><table><thead><tr><th>Placa</th><th style="text-align:right;">Tempo acumulado</th></tr></thead><tbody>
    """

    if veiculos_maior_tempo:
        for item in veiculos_maior_tempo:
            html += f"<tr><td>{escape_html(item['placa'])}</td><td class='td-num'>{escape_html(item['tempo_formatado'])}</td></tr>"
    else:
        html += "<tr><td colspan='2' class='empty'>Nenhum dado encontrado</td></tr>"

    html += f"""</tbody></table></div>
              <div class="footer-note">O comparativo com período anterior só aparece quando data inicial e final são informadas.</div>
            </div>
          </div>
        </div>
      </body>
    </html>
    """
    return html


# =========================
# PAINEL WEB — ORDENS DE SERVIÇO
# =========================

@app.get("/painel/ordens-servico", response_class=HTMLResponse)
def painel_ordens_servico(
    request: Request,
    placa: str = "",
    matricula: str = "",
    mecanico: str = "",
    km: str = "",
    servico: str = "",
    data_cadastro: str = "",
    db: Session = Depends(get_db),
):
    if not usuario_logado(request):
        return RedirectResponse(url="/login", status_code=303)

    rows = buscar_linhas_ordens(db, placa=placa, matricula=matricula, mecanico=mecanico, km=km, servico=servico, data_cadastro=data_cadastro)
    filtro_query = f"placa={placa}&matricula={matricula}&mecanico={mecanico}&km={km}&servico={servico}&data_cadastro={data_cadastro}"

    html = f"""
    <html>
      <head>
        <meta charset="UTF-8">
        <title>Painel de Ordens de Serviço</title>
        <style>
          body {{ font-family: Arial, sans-serif; margin: 20px; background: #f7f7f7; }}
          h1 {{ margin-bottom: 20px; }}
          .card {{ background: #fff; border-radius: 12px; padding: 16px; box-shadow: 0 2px 10px rgba(0,0,0,0.08); overflow-x: auto; margin-bottom: 20px; }}
          .filtros {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin-bottom: 12px; }}
          .filtros input {{ padding: 10px; border: 1px solid #ccc; border-radius: 8px; font-size: 14px; }}
          .acoes {{ display: flex; gap: 10px; flex-wrap: wrap; margin-top: 12px; }}
          .btn {{ padding: 10px 16px; border: none; border-radius: 8px; cursor: pointer; font-size: 14px; text-decoration: none; display: inline-block; }}
          .btn-buscar {{ background: #1f2937; color: white; }}
          .btn-limpar {{ background: #e5e7eb; color: #111827; }}
          .btn-exportar {{ background: #0f766e; color: white; }}
          .btn-excluir {{ background: #dc2626; color: white; }}
          table {{ border-collapse: collapse; width: 100%; background: #fff; }}
          th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; vertical-align: top; white-space: nowrap; }}
          th {{ background: #1f2937; color: white; }}
          tr:nth-child(even) {{ background: #fafafa; }}
          tr:hover {{ background: #f0f7ff; }}
          .total {{ margin-bottom: 10px; font-weight: bold; }}
          form.inline {{ display: inline; }}
          .produtos {{ font-size: 12px; color: #374151; margin-top: 4px; white-space: normal; }}
          .foto-thumb {{ height: 48px; border-radius: 6px; margin-top: 4px; display: block; cursor: pointer; }}
        </style>
      </head>
      <body>
        <h1>Painel de Ordens de Serviço</h1>
        <div class="card">
          <div class="acoes" style="margin-bottom:12px;">
            <a href="/painel/dashboard" class="btn btn-buscar">Ir para dashboard</a>
            <a href="/painel/produtos" class="btn btn-limpar">Ir para produtos</a>
            <a href="/painel/mecanicos" class="btn btn-limpar">Ir para mecânicos</a>
            <a href="/painel/servicos" class="btn btn-limpar">Ir para serviços</a>
            <a href="/logout" class="btn btn-limpar">Sair</a>
          </div>
          <form method="get" action="/painel/ordens-servico">
            <div class="filtros">
              <input type="text" name="placa" placeholder="Filtrar por placa" value="{escape_html(placa)}">
              <input type="text" name="matricula" placeholder="Filtrar por matrícula" value="{escape_html(matricula)}">
              <input type="text" name="mecanico" placeholder="Filtrar por mecânico" value="{escape_html(mecanico)}">
              <input type="text" name="km" placeholder="Filtrar por KM" value="{escape_html(km)}">
              <input type="text" name="servico" placeholder="Filtrar por serviço" value="{escape_html(servico)}">
              <input type="text" name="data_cadastro" placeholder="Filtrar por data" value="{escape_html(data_cadastro)}">
            </div>
            <div class="acoes">
              <button type="submit" class="btn btn-buscar">Buscar</button>
              <a href="/painel/ordens-servico" class="btn btn-limpar">Limpar filtros</a>
              <a href="/painel/ordens-servico/exportar/json?{filtro_query}" class="btn btn-exportar">Exportar JSON</a>
              <a href="/painel/ordens-servico/exportar/xlsx?{filtro_query}" class="btn btn-exportar">Exportar XLSX</a>
            </div>
          </form>
        </div>
        <div class="card">
          <div class="total">Total de linhas: {len(rows)}</div>
          <table>
            <thead>
              <tr>
                <th>ID</th><th>Placa</th><th>Matrícula</th><th>Mecânico</th><th>KM</th>
                <th>Hora início</th><th>Hora fim</th><th>Serviço / Produtos</th>
                <th>Foto</th><th>Observações</th><th>Status FrotaWeb</th><th>Nº O.S. FrotaWeb</th><th>Retorno FrotaWeb</th><th>Data cadastro</th><th>Ações</th>
              </tr>
            </thead>
            <tbody>
    """

    if rows:
        ids_exibidos = set()

        for row in rows:
            observacoes = row.observacoes if row.observacoes else "-"
            ordem = db.query(OrdemServicoModel).filter(OrdemServicoModel.id == row.id).first()

            primeira_linha_da_ordem = row.id not in ids_exibidos

            produtos_html = ""
            if ordem and ordem.produtos_consumidos:
                linhas = [
                    f"{escape_html(p.produto_nome)} ({escape_html(p.quantidade)})"
                    for p in ordem.produtos_consumidos
                ]
                produtos_html = (
                    "<div class='produtos'><strong>Produtos:</strong> "
                    + ", ".join(linhas)
                    + "</div>"
                )

            foto_html = ""
            if primeira_linha_da_ordem:
                if row.foto:
                    foto_html = (
                        f"<a href='{escape_html(row.foto)}' target='_blank'>"
                        f"<img src='{escape_html(row.foto)}' class='foto-thumb' alt='foto'>"
                        f"</a>"
                    )
                else:
                    foto_html = "<span style='color:#9ca3af;font-size:12px;'>Sem foto</span>"

            botao_excluir = ""
            if primeira_linha_da_ordem:
                botao_enviar_frotaweb = ""
                if (ordem.frotaweb_status or "PENDENTE") != "ENVIADA":
                    botao_enviar_frotaweb = f"""
                    <form class="inline" method="post" action="/painel/ordens-servico/{row.id}/enviar-frotaweb"
                          onsubmit="return confirm('Enviar a O.S. {row.id} para o FrotaWeb agora?');">
                      <button type="submit" class="btn btn-exportar">Enviar FrotaWeb</button>
                    </form>
                    """

                botao_excluir = f"""
                {botao_enviar_frotaweb}
                <form class="inline" method="post" action="/ordens-servico/{row.id}/excluir"
                      onsubmit="return confirm('Deseja excluir a O.S. {row.id}?');">
                  <button type="submit" class="btn btn-excluir">Excluir</button>
                </form>
                """
                ids_exibidos.add(row.id)

                html += f"""
                  <tr>
                    <td>{row.id}</td>
                    <td>{escape_html(row.placa)}</td>
                    <td>{escape_html(row.matricula)}</td>
                    <td>{escape_html(row.mecanico)}</td>
                    <td>{escape_html(row.km)}</td>
                    <td>{escape_html(row.hora_inicio)}</td>
                    <td>{escape_html(row.hora_fim)}</td>
                    <td>{escape_html(row.servico)}{produtos_html}</td>
                    <td>{foto_html}</td>
                    <td>{escape_html(observacoes)}</td>
                    <td>{escape_html(ordem.frotaweb_status if ordem else getattr(row, "frotaweb_status", "PENDENTE"))}</td>
                    <td>{escape_html(ordem.frotaweb_numero_os if ordem else getattr(row, "frotaweb_numero_os", ""))}</td>
                    <td style="white-space:normal;max-width:320px;">{escape_html((ordem.frotaweb_retorno if ordem else getattr(row, "frotaweb_retorno", "")) or "")[:500]}</td>
                    <td>{escape_html(row.data_cadastro)}</td>
                    <td>{botao_excluir}</td>
                  </tr>
                """
    else:
        html += "<tr><td colspan='15' style='text-align:center;'>Nenhum registro encontrado</td></tr>"

    html += "</tbody></table></div></body></html>"
    return html


# =========================
# EXPORTAÇÕES
# =========================

@app.get("/painel/ordens-servico/exportar/json")
def exportar_ordens_json(
    placa: str = "", matricula: str = "", mecanico: str = "",
    km: str = "", servico: str = "", data_cadastro: str = "",
    db: Session = Depends(get_db),
):
    rows = buscar_linhas_ordens(db, placa=placa, matricula=matricula, mecanico=mecanico, km=km, servico=servico, data_cadastro=data_cadastro)
    return JSONResponse(content=[
        {
            "id": row.id, "placa": row.placa, "matricula": row.matricula,
            "mecanico": row.mecanico, "km": row.km, "hora_inicio": row.hora_inicio,
            "hora_fim": row.hora_fim, "servico": row.servico,
            "observacoes": row.observacoes, "data_cadastro": row.data_cadastro,
            "foto_url": row.foto,
            "frotaweb_status": getattr(row, "frotaweb_status", "PENDENTE"),
            "frotaweb_numero_os": getattr(row, "frotaweb_numero_os", ""),
            "frotaweb_retorno": getattr(row, "frotaweb_retorno", ""),
        }
        for row in rows
    ])


@app.get("/painel/ordens-servico/exportar/xlsx")
def exportar_ordens_xlsx(
    placa: str = "", matricula: str = "", mecanico: str = "",
    km: str = "", servico: str = "", data_cadastro: str = "",
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Pacote openpyxl não instalado no servidor")

    rows = buscar_linhas_ordens(db, placa=placa, matricula=matricula, mecanico=mecanico, km=km, servico=servico, data_cadastro=data_cadastro)

    wb = Workbook()
    ws = wb.active
    ws.title = "OrdensServico"
    ws.append(["ID", "Placa", "Matrícula", "Mecânico", "KM", "Hora início", "Hora fim", "Serviço", "Observações", "Status FrotaWeb", "Nº O.S. FrotaWeb", "Retorno FrotaWeb", "Data cadastro", "Foto URL"])

    for row in rows:
        ws.append([
            row.id, row.placa, row.matricula, row.mecanico, row.km,
            row.hora_inicio, row.hora_fim, row.servico, row.observacoes,
            getattr(row, "frotaweb_status", "PENDENTE"), getattr(row, "frotaweb_numero_os", ""), getattr(row, "frotaweb_retorno", ""),
            row.data_cadastro, row.foto or "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ordens_servico.xlsx"},
    )
